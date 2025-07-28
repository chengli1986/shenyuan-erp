"""
修复手动更新后的Excel文件
1. 重新排序序列号
2. 修复合价公式为单价*数量
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def fix_updated_dual_system_file():
    """修复手动更新后的双系统Excel文件"""
    
    file_path = os.path.join(os.path.dirname(__file__), 'templates', '合同清单-完整版含两个系统.xlsx')
    
    if not os.path.exists(file_path):
        print(f"[ERROR] 文件不存在: {file_path}")
        return None
    
    try:
        # 加载工作簿
        wb = load_workbook(file_path)
        
        print(f"[INFO] 修复双系统Excel文件...")
        print(f"[INFO] 文件: {os.path.basename(file_path)}")
        print(f"[INFO] 工作表: {wb.sheetnames}")
        
        # 处理每个工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n[INFO] 处理工作表: {sheet_name}")
            
            # 找到数据范围
            data_start_row = 4  # 数据从第4行开始
            data_end_row = ws.max_row
            
            # 找到实际数据结束行（排除合计行）
            for row in range(data_start_row, ws.max_row + 1):
                device_name = ws.cell(row=row, column=2).value
                if not device_name or str(device_name).strip() in ['合计', '总计', '小计', '']:
                    data_end_row = row - 1
                    break
            
            print(f"  - 数据行范围: {data_start_row} 到 {data_end_row}")
            
            # 1. 重新排序序列号
            sequence_count = 0
            print(f"  - 重新排序序列号...")
            
            for row in range(data_start_row, data_end_row + 1):
                device_name = ws.cell(row=row, column=2).value
                if device_name and str(device_name).strip():
                    sequence_count += 1
                    ws.cell(row=row, column=1, value=sequence_count)  # A列：序号
                    
            print(f"    共更新 {sequence_count} 个序列号")
            
            # 2. 修复合价公式
            formula_count = 0
            print(f"  - 修复合价公式...")
            
            for row in range(data_start_row, data_end_row + 1):
                device_name = ws.cell(row=row, column=2).value
                if device_name and str(device_name).strip():
                    # 设置合价公式：=G行号*F行号（单价×数量）
                    formula = f"=G{row}*F{row}"
                    ws.cell(row=row, column=8, value=formula)  # H列：合价
                    formula_count += 1
                    
                    print(f"    第{row}行: {str(device_name)[:20]:<20} 公式: {formula}")
            
            print(f"    共设置 {formula_count} 个公式")
            
            # 3. 更新合计公式
            total_row = None
            for row in range(data_end_row + 1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and '合计' in str(cell_value):
                    total_row = row
                    break
            
            if total_row:
                # 更新合计公式
                sum_formula = f"=SUM(H{data_start_row}:H{data_end_row})"
                ws.cell(row=total_row, column=8, value=sum_formula)
                print(f"  - 更新合计公式: {sum_formula}")
            else:
                print(f"  - [WARNING] 未找到合计行")
            
            # 4. 设置单元格格式
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in range(data_start_row, data_end_row + 1):
                for col in range(1, 10):
                    cell = ws.cell(row=row, column=col)
                    if cell.value:  # 只处理有内容的单元格
                        cell.border = thin_border
                        
                        # 设置对齐方式
                        if col in [1, 5, 6, 7, 8]:  # 序号、单位、数量、单价、合价
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                        
                        # 设置数字格式
                        if col == 6:  # 数量
                            cell.number_format = '0.00'
                        elif col in [7, 8]:  # 单价、合价
                            cell.number_format = '#,##0.00'
        
        # 保存文件
        output_path = os.path.join(os.path.dirname(__file__), 'templates', '合同清单-完整版含两个系统-修复版.xlsx')
        wb.save(output_path)
        
        print(f"\n[OK] 文件修复完成!")
        print(f"[INFO] 输出文件: {output_path}")
        print(f"[INFO] 修复内容:")
        print(f"  - 重新排序了所有序列号")
        print(f"  - 修复了所有合价公式为 =G*F 格式")
        print(f"  - 更新了合计公式")
        print(f"  - 统一了单元格格式")
        
        return output_path
        
    except Exception as e:
        print(f"[ERROR] 修复失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def verify_fixed_file():
    """验证修复后的文件"""
    file_path = os.path.join(os.path.dirname(__file__), 'templates', '合同清单-完整版含两个系统-修复版.xlsx')
    
    if not os.path.exists(file_path):
        print("[ERROR] 修复后的文件不存在")
        return
    
    try:
        wb = load_workbook(file_path)
        
        print(f"\n[INFO] 验证修复后的文件:")
        print(f"  - 文件名: {os.path.basename(file_path)}")
        print(f"  - 工作表数: {len(wb.sheetnames)}")
        
        # 验证每个工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n[INFO] 工作表: {sheet_name}")
            
            # 检查序列号
            sequence_numbers = []
            formulas = []
            
            for row in range(4, 20):  # 检查前几行
                device_name = ws.cell(row=row, column=2).value
                if device_name and str(device_name).strip():
                    seq_num = ws.cell(row=row, column=1).value
                    formula = ws.cell(row=row, column=8).value
                    sequence_numbers.append(seq_num)
                    formulas.append(formula)
                    
                    if len(sequence_numbers) <= 5:  # 只显示前5行
                        print(f"  第{row}行: 序号:{seq_num} 设备:{str(device_name)[:20]:<20} 公式:{formula}")
            
            print(f"  - 序列号范围: 1 到 {max(sequence_numbers) if sequence_numbers else 0}")
            print(f"  - 公式检查: {len([f for f in formulas if f and str(f).startswith('=G')]) if formulas else 0} 个正确公式")
            
            # 检查合计公式
            for row in range(ws.max_row - 10, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and '合计' in str(cell_value):
                    sum_formula = ws.cell(row=row, column=8).value
                    print(f"  - 合计公式: {sum_formula}")
                    break
        
        print(f"\n[SUCCESS] 验证完成!")
        
    except Exception as e:
        print(f"[ERROR] 验证失败: {str(e)}")

def main():
    """主函数"""
    print("="*80)
    print("修复手动更新的双系统Excel文件")
    print("="*80)
    
    # 修复文件
    result = fix_updated_dual_system_file()
    
    if result:
        # 验证修复结果
        verify_fixed_file()
        
        print(f"\n[SUCCESS] 修复完成!")
        print(f"输出文件: {result}")
        print("\n修复说明:")
        print("1. 序列号重新排序：从1开始连续编号")
        print("2. 合价公式修复：统一为 =G*F 格式（单价×数量）")
        print("3. 合计公式更新：使用SUM函数汇总")
        print("4. 格式统一：边框、对齐、数字格式")
        
        print(f"\n使用说明:")
        print("- 原文件保持不变")
        print("- 新文件添加了'-修复版'后缀")
        print("- 所有计算使用Excel公式，动态更新")
        print("- 可以直接使用或替换原文件")
    else:
        print("[FAILED] 修复失败")

if __name__ == "__main__":
    main()