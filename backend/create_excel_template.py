"""
创建标准的合同清单Excel模板
"""

import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_contract_template():
    """创建合同清单Excel模板"""
    
    # 创建工作簿
    wb = Workbook()
    
    # 定义标准列名和说明
    columns_info = [
        ('序号', '数字序号，从1开始'),
        ('设备名称', '设备或材料的名称，必填'),
        ('品牌型号', '品牌和型号信息'),
        ('技术规格参数', '详细的技术参数说明'),
        ('单位', '计量单位，如：台、个、套、米等'),
        ('数量', '数量，必须是数字，必填'),
        ('综合单价', '单价（元），必须是数字，必填'),
        ('合价', '总价（元），可以使用公式：数量×单价'),
        ('原产地', '产品原产地，默认为"中国"'),
        ('备注', '其他说明信息')
    ]
    
    # 示例数据
    sample_data = [
        {
            '序号': 1,
            '设备名称': '网络摄像机',
            '品牌型号': '海康威视 DS-2CD2345',
            '技术规格参数': '400万像素，H.265编码，支持POE供电，红外夜视50米',
            '单位': '台',
            '数量': 20,
            '综合单价': 1200,
            '合价': '=F2*G2',
            '原产地': '中国',
            '备注': '室外防水型'
        },
        {
            '序号': 2,
            '设备名称': '网络硬盘录像机',
            '品牌型号': '海康威视 DS-7832N-K2',
            '技术规格参数': '32路网络录像机，支持4K，8盘位，支持RAID',
            '单位': '台',
            '数量': 1,
            '综合单价': 4500,
            '合价': '=F3*G3',
            '原产地': '中国',
            '备注': '含8块4TB监控硬盘'
        },
        {
            '序号': 3,
            '设备名称': '千兆交换机',
            '品牌型号': '华为 S5720-28P-LI-AC',
            '技术规格参数': '24口千兆电+4千兆光，支持POE+，370W',
            '单位': '台',
            '数量': 5,
            '综合单价': 3500,
            '合价': '=F4*G4',
            '原产地': '中国',
            '备注': '核心机房使用'
        }
    ]
    
    # 创建多个系统的工作表
    systems = [
        ('视频监控', '视频监控系统设备清单'),
        ('门禁系统', '门禁对讲系统设备清单'),
        ('综合布线', '综合布线系统材料清单')
    ]
    
    # 删除默认的Sheet
    wb.remove(wb.active)
    
    # 样式定义
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 为每个系统创建工作表
    for sheet_name, sheet_title in systems:
        ws = wb.create_sheet(title=sheet_name)
        
        # 写入标题
        ws.merge_cells('A1:J1')
        ws['A1'] = sheet_title
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入列说明
        ws.merge_cells('A2:J2')
        ws['A2'] = '说明：请按照以下格式填写设备清单，带*的为必填项'
        ws['A2'].font = Font(italic=True, size=10, color="FF0000")
        
        # 写入表头
        for col_idx, (col_name, col_desc) in enumerate(columns_info, 1):
            cell = ws.cell(row=3, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
            # 设置列宽
            col_letter = get_column_letter(col_idx)
            if col_idx == 4:  # 技术规格参数列
                ws.column_dimensions[col_letter].width = 40
            elif col_idx in [2, 3]:  # 设备名称和品牌型号
                ws.column_dimensions[col_letter].width = 25
            elif col_idx == 10:  # 备注列
                ws.column_dimensions[col_letter].width = 20
            else:
                ws.column_dimensions[col_letter].width = 12
        
        # 如果是第一个工作表，写入示例数据
        if sheet_name == '视频监控':
            for row_idx, data in enumerate(sample_data, 4):
                for col_idx, col_name in enumerate([col[0] for col in columns_info], 1):
                    value = data.get(col_name, '')
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    
                    # 对齐方式
                    if col_idx in [1, 5, 6, 7, 8]:  # 数字列
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # 添加合计行
        last_row = 7 if sheet_name == '视频监控' else 4
        ws.merge_cells(f'A{last_row}:E{last_row}')
        ws[f'A{last_row}'] = '合计'
        ws[f'A{last_row}'].font = Font(bold=True)
        ws[f'A{last_row}'].alignment = Alignment(horizontal="right", vertical="center")
        
        # 设置合计公式
        if sheet_name == '视频监控':
            ws[f'F{last_row}'] = '=SUM(F4:F6)'
            ws[f'H{last_row}'] = '=SUM(H4:H6)'
        
        # 添加页脚说明
        footer_row = last_row + 2
        ws.merge_cells(f'A{footer_row}:J{footer_row}')
        ws[f'A{footer_row}'] = '填写说明：1.设备名称、数量、综合单价为必填项；2.合价会自动计算；3.每个系统请在对应的工作表中填写'
        ws[f'A{footer_row}'].font = Font(size=9, color="666666")
    
    # 保存文件
    template_dir = os.path.join(os.path.dirname(__file__), 'templates')
    os.makedirs(template_dir, exist_ok=True)
    
    file_path = os.path.join(template_dir, '合同清单模板.xlsx')
    wb.save(file_path)
    
    print(f"[OK] Excel模板创建成功: {file_path}")
    
    # 创建填写指南
    guide_content = """# 合同清单Excel填写指南

## 1. 文件结构
- 每个系统使用独立的工作表（Sheet）
- 工作表名称即为系统名称（如：视频监控、门禁系统等）

## 2. 必填字段
- **设备名称**：设备或材料的名称
- **数量**：必须是数字
- **综合单价**：单价（元），必须是数字

## 3. 列名规范
请使用以下标准列名：
- 序号
- 设备名称
- 品牌型号
- 技术规格参数
- 单位
- 数量
- 综合单价
- 合价
- 原产地
- 备注

## 4. 注意事项
1. 不要修改列名
2. 不要使用合并单元格（标题行除外）
3. 确保数字字段不包含文字
4. 每个系统单独一个工作表
5. 删除示例数据后再填写真实数据

## 5. 常见问题
- 如果某些信息暂时没有，可以留空
- 合价可以使用Excel公式自动计算
- 原产地默认为"中国"
"""
    
    guide_path = os.path.join(template_dir, '填写指南.md')
    with open(guide_path, 'w', encoding='utf-8') as f:
        f.write(guide_content)
    
    print(f"[OK] 填写指南创建成功: {guide_path}")

if __name__ == "__main__":
    create_contract_template()