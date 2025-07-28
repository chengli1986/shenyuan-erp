"""
修复Excel文件的列错位问题
确保数据在正确的列中：单位、数量、单价、合价
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def analyze_current_structure():
    """分析当前文件的结构，识别错位问题"""
    file_path = os.path.join(os.path.dirname(__file__), 'templates', '合同清单-视频监控完整内容-公式版.xlsx')
    
    if not os.path.exists(file_path):
        print(f"[ERROR] 文件不存在: {file_path}")
        return None
    
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        print(f"[INFO] 分析当前文件结构...")
        
        # 显示表头
        print(f"\n[INFO] 当前表头 (第3行):")
        headers = []
        for col in range(1, 10):
            header_value = ws.cell(row=3, column=col).value
            if header_value:
                headers.append(str(header_value))
                print(f"  {chr(64+col)}列: {header_value}")
        
        # 显示前3行数据，分析错位情况
        print(f"\n[INFO] 前3行数据分析:")
        for row in range(4, 7):
            print(f"\n  第{row}行:")
            for col in range(1, 10):
                cell_value = ws.cell(row=row, column=col).value
                col_letter = chr(64 + col)
                header = headers[col-1] if col-1 < len(headers) else "未知"
                print(f"    {col_letter}列({header}): {cell_value}")
        
        return {
            'workbook': wb,
            'worksheet': ws,
            'headers': headers
        }
        
    except Exception as e:
        print(f"[ERROR] 分析失败: {str(e)}")
        return None

def fix_column_alignment():
    """修复列错位问题"""
    
    # 先分析当前结构
    analysis = analyze_current_structure()
    if not analysis:
        return None
    
    ws = analysis['worksheet']
    wb = analysis['workbook']
    
    print(f"\n[INFO] 开始修复列错位...")
    
    # 标准列头定义
    standard_headers = ['序号', '设备名称', '品牌型号', '技术规格参数', '单位', '数量', '综合单价', '合价', '备注']
    
    # 找到数据范围
    data_start_row = 4
    data_end_row = ws.max_row
    
    # 找到实际数据结束行
    for row in range(data_start_row, ws.max_row + 1):
        device_name = ws.cell(row=row, column=2).value
        if not device_name or str(device_name).strip() in ['合计', '总计', '小计', '']:
            data_end_row = row - 1
            break
    
    print(f"[INFO] 数据行范围: {data_start_row} 到 {data_end_row}")
    
    # 备份原始数据
    original_data = []
    for row in range(data_start_row, data_end_row + 1):
        row_data = []
        for col in range(1, 10):
            cell_value = ws.cell(row=row, column=col).value
            row_data.append(cell_value)
        original_data.append(row_data)
    
    print(f"[INFO] 已备份 {len(original_data)} 行原始数据")
    
    # 分析并重新排列数据
    fixed_data = []
    
    for i, row_data in enumerate(original_data):
        current_row = data_start_row + i
        
        # 原始数据（根据观察到的错位情况）
        serial_num = row_data[0] if row_data[0] else (i + 1)  # A列：序号
        device_name = row_data[1] if row_data[1] else ''      # B列：设备名称
        brand_model = row_data[2] if row_data[2] else ''      # C列：品牌型号
        specification = row_data[3] if row_data[3] else ''    # D列：技术规格参数
        
        # 错位问题修复：
        # 根据您的描述：单位显示金额，数量显示单位，单价显示数量
        # 这意味着数据向右偏移了
        
        # 尝试从原始数据中正确提取信息
        unit = None
        quantity = None
        unit_price = None
        
        # 寻找单位（应该是文本，如"台"、"个"、"套"等）
        for j in range(4, 8):
            if j < len(row_data) and row_data[j]:
                val = str(row_data[j]).strip()
                if val in ['台', '个', '套', '对', '块', '根', '米', '箱', '卷', '张', '片', '只', '条', '组']:
                    unit = val
                    unit_col = j
                    break
        
        # 寻找数量（应该是数字，且相对较小）
        for j in range(4, 8):
            if j < len(row_data) and row_data[j] is not None:
                try:
                    val = float(row_data[j])
                    if 0 < val < 10000:  # 合理的数量范围
                        quantity = val
                        quantity_col = j
                        break
                except:
                    continue
        
        # 寻找单价（应该是数字，且相对较大）
        for j in range(4, 8):
            if j < len(row_data) and row_data[j] is not None:
                try:
                    val = float(row_data[j])
                    if val > 1:  # 单价应该大于1
                        unit_price = val
                        unit_price_col = j
                        break
                except:
                    continue
        
        # 默认值处理
        if not unit:
            unit = '台'
        if not quantity:
            quantity = 1
        if not unit_price:
            unit_price = 0
        
        remarks = row_data[8] if len(row_data) > 8 and row_data[8] else ''
        
        # 重新排列的数据
        fixed_row = [
            i + 1,              # A列：序号（重新编号）
            device_name,        # B列：设备名称
            brand_model,        # C列：品牌型号
            specification,      # D列：技术规格参数
            unit,              # E列：单位
            quantity,          # F列：数量
            unit_price,        # G列：综合单价
            f"=G{current_row}*F{current_row}",  # H列：合价公式
            remarks            # I列：备注
        ]
        
        fixed_data.append(fixed_row)
        
        print(f"  第{current_row}行: {device_name[:20]:<20} 单位:{unit} 数量:{quantity} 单价:{unit_price}")
    
    # 清空原有数据区域
    for row in range(data_start_row, data_end_row + 1):
        for col in range(1, 10):
            ws.cell(row=row, column=col).value = None
    
    # 写入修复后的数据
    print(f"\n[INFO] 写入修复后的数据...")
    for i, row_data in enumerate(fixed_data):
        current_row = data_start_row + i
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = value
            
            # 设置单元格格式
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border
            
            # 设置对齐方式
            if col in [1, 5, 6, 7, 8]:  # 序号、单位、数量、单价、合价
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # 设置数字格式
            if col in [6, 7, 8]:  # 数量、单价、合价
                if col == 6:  # 数量
                    cell.number_format = '0.00'
                else:  # 单价、合价
                    cell.number_format = '#,##0.00'
    
    # 添加合计行
    total_row = data_start_row + len(fixed_data) + 1
    
    # 样式定义
    total_font = Font(bold=True, size=12)
    total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    total_alignment = Alignment(horizontal="center", vertical="center")
    amount_alignment = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 合并单元格并添加"合计"
    ws.merge_cells(f'A{total_row}:F{total_row}')
    total_cell = ws[f'A{total_row}']
    total_cell.value = '合计'
    total_cell.font = total_font
    total_cell.fill = total_fill
    total_cell.alignment = total_alignment
    total_cell.border = thin_border
    
    # 添加合计公式
    sum_formula = f"=SUM(H{data_start_row}:H{data_start_row + len(fixed_data) - 1})"
    amount_cell = ws[f'H{total_row}']
    amount_cell.value = sum_formula
    amount_cell.font = total_font
    amount_cell.fill = total_fill
    amount_cell.alignment = amount_alignment
    amount_cell.border = thin_border
    amount_cell.number_format = '#,##0.00'
    
    # G列和I列样式
    for col in ['G', 'I']:
        cell = ws[f'{col}{total_row}']
        cell.border = thin_border
        cell.fill = total_fill
    
    # 保存文件
    output_path = os.path.join(os.path.dirname(__file__), 'templates', '合同清单-视频监控完整内容-修复列错位.xlsx')
    wb.save(output_path)
    
    print(f"\n[OK] 列错位修复完成!")
    print(f"[INFO] 输出文件: {output_path}")
    print(f"[INFO] 修复设备数: {len(fixed_data)}")
    print(f"[INFO] 合计公式: {sum_formula}")
    
    return output_path

def verify_fixed_file():
    """验证修复后的文件"""
    file_path = os.path.join(os.path.dirname(__file__), 'templates', '合同清单-视频监控完整内容-修复列错位.xlsx')
    
    if not os.path.exists(file_path):
        print("[ERROR] 修复后的文件不存在")
        return
    
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        print(f"\n[INFO] 验证修复后的文件:")
        
        # 验证表头
        print(f"\n[INFO] 标准表头验证:")
        standard_headers = ['序号', '设备名称', '品牌型号', '技术规格参数', '单位', '数量', '综合单价', '合价', '备注']
        for col, expected_header in enumerate(standard_headers, 1):
            actual_header = ws.cell(row=3, column=col).value
            status = "✓" if str(actual_header) == expected_header else "✗"
            print(f"  {chr(64+col)}列: {expected_header} {status}")
        
        # 验证前3行数据
        print(f"\n[INFO] 前3行数据验证:")
        for row in range(4, 7):
            serial_num = ws.cell(row=row, column=1).value
            device_name = ws.cell(row=row, column=2).value
            unit = ws.cell(row=row, column=5).value
            quantity = ws.cell(row=row, column=6).value
            unit_price = ws.cell(row=row, column=7).value
            formula = ws.cell(row=row, column=8).value
            
            print(f"  第{row}行: 序号:{serial_num} 设备:{str(device_name)[:15]:<15} 单位:{unit} 数量:{quantity} 单价:{unit_price} 公式:{formula}")
        
        print(f"\n[SUCCESS] 文件验证完成!")
        
    except Exception as e:
        print(f"[ERROR] 验证失败: {str(e)}")

def main():
    """主函数"""
    print("="*80)
    print("修复Excel列错位问题")
    print("="*80)
    
    # 修复列错位
    result = fix_column_alignment()
    
    if result:
        # 验证修复结果
        verify_fixed_file()
        
        print(f"\n[SUCCESS] 列错位修复完成!")
        print(f"输出文件: {result}")
        print("\n修复内容:")
        print("- 重新排列了所有列的数据")
        print("- 确保单位在E列，数量在F列，单价在G列")
        print("- 修复了公式引用（=G*F）")
        print("- 重新编号了序列号")
        print("- 更新了合计公式")
        
        print(f"\n列结构说明:")
        print("A列：序号    B列：设备名称    C列：品牌型号    D列：技术规格参数")
        print("E列：单位    F列：数量        G列：综合单价    H列：合价")
        print("I列：备注")
    else:
        print("[FAILED] 修复失败")

if __name__ == "__main__":
    main()