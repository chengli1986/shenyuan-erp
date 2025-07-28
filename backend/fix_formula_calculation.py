"""
修复Excel文件，保留公式计算而不是直接填入金额
使用单价×数量的公式格式
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def fix_formula_calculation():
    """修复公式计算，保留=G*F格式而不是直接金额"""
    
    # 文件路径
    file_path = os.path.join(os.path.dirname(__file__), 'templates', '合同清单-视频监控完整内容-最终版.xlsx')
    
    if not os.path.exists(file_path):
        print(f"[ERROR] 文件不存在: {file_path}")
        return None
    
    try:
        # 加载工作簿
        wb = load_workbook(file_path)
        ws = wb.active
        
        print(f"[INFO] 正在修复公式计算...")
        print(f"[INFO] 文件: {os.path.basename(file_path)}")
        
        # 找到数据范围
        data_start_row = 4
        data_end_row = ws.max_row
        
        # 找到实际的数据结束行
        for row in range(data_start_row, ws.max_row + 1):
            device_name = ws.cell(row=row, column=2).value
            if not device_name or str(device_name).strip() in ['合计', '总计', '小计', '']:
                data_end_row = row - 1
                break
        
        print(f"[INFO] 数据行范围: {data_start_row} 到 {data_end_row}")
        
        # 修复每一行的公式
        device_count = 0
        formula_count = 0
        
        print(f"\n[INFO] 修复公式:")
        for row in range(data_start_row, data_end_row + 1):
            device_name = ws.cell(row=row, column=2).value
            quantity = ws.cell(row=row, column=6).value  # F列：数量
            unit_price = ws.cell(row=row, column=7).value  # G列：单价
            total_price_cell = ws.cell(row=row, column=8)  # H列：合价
            
            if device_name and str(device_name).strip():
                device_count += 1
                
                # 设置公式：=G行号*F行号（单价×数量）
                formula = f"=G{row}*F{row}"
                total_price_cell.value = formula
                formula_count += 1
                
                print(f"  第{row}行: {str(device_name)[:25]:<25} 公式: {formula}")
        
        print(f"\n[INFO] 公式修复结果:")
        print(f"  - 处理设备数: {device_count}")
        print(f"  - 设置公式数: {formula_count}")
        
        # 添加合计行
        total_row = data_end_row + 2
        
        # 样式定义
        total_font = Font(bold=True, size=12)
        total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        total_alignment = Alignment(horizontal="center", vertical="center")
        amount_alignment = Alignment(horizontal="right", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 合并单元格并添加"合计"
        ws.merge_cells(f'A{total_row}:F{total_row}')
        total_cell = ws[f'A{total_row}']
        total_cell.value = '合计'
        total_cell.font = total_font
        total_cell.fill = total_fill
        total_cell.alignment = total_alignment
        total_cell.border = thin_border
        
        # 添加合计公式（SUM函数）
        sum_formula = f"=SUM(H{data_start_row}:H{data_end_row})"
        amount_cell = ws[f'H{total_row}']
        amount_cell.value = sum_formula
        amount_cell.font = total_font
        amount_cell.fill = total_fill
        amount_cell.alignment = amount_alignment
        amount_cell.border = thin_border
        amount_cell.number_format = '#,##0.00'
        
        print(f"[INFO] 合计公式: {sum_formula}")
        
        # G列和I列也加上边框和样式
        for col in ['G', 'I']:
            cell = ws[f'{col}{total_row}']
            cell.border = thin_border
            cell.fill = total_fill
        
        # 添加统计摘要
        summary_row = total_row + 2
        ws.merge_cells(f'A{summary_row}:I{summary_row}')
        summary_cell = ws[f'A{summary_row}']
        summary_cell.value = f'项目摘要：视频监控系统共 {device_count} 项设备，所有金额均使用公式自动计算（单价×数量）'
        summary_cell.font = Font(size=11, color="333333")
        summary_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # 添加公式说明
        formula_row = summary_row + 1
        ws.merge_cells(f'A{formula_row}:I{formula_row}')
        formula_cell = ws[f'A{formula_row}']
        formula_cell.value = f'计算说明：合价 = 单价 × 数量，总计使用SUM函数汇总，所有计算自动更新'
        formula_cell.font = Font(size=10, italic=True, color="666666")
        formula_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # 保存文件
        output_path = os.path.join(os.path.dirname(__file__), 'templates', '合同清单-视频监控完整内容-公式版.xlsx')
        wb.save(output_path)
        
        print(f"\n[OK] 公式修复完成!")
        print(f"[INFO] 输出文件: {output_path}")
        print(f"[INFO] 设备总数: {device_count} 项")
        print(f"[INFO] 公式设置: {formula_count} 个")
        print(f"[INFO] 合计公式: {sum_formula}")
        
        return {
            'file_path': output_path,
            'device_count': device_count,
            'formula_count': formula_count,
            'sum_formula': sum_formula
        }
        
    except Exception as e:
        print(f"[ERROR] 修复失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def verify_formulas():
    """验证公式是否正确设置"""
    file_path = os.path.join(os.path.dirname(__file__), 'templates', '合同清单-视频监控完整内容-公式版.xlsx')
    
    if not os.path.exists(file_path):
        print("[ERROR] 公式版文件不存在")
        return
    
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        print(f"\n[INFO] 验证公式设置:")
        print(f"  - 工作表: {ws.title}")
        
        # 检查前几行的公式
        print(f"\n[INFO] 前5行公式检查:")
        for row in range(4, 9):
            device_name = ws.cell(row=row, column=2).value
            quantity = ws.cell(row=row, column=6).value
            unit_price = ws.cell(row=row, column=7).value
            formula = ws.cell(row=row, column=8).value
            
            if device_name:
                print(f"  第{row}行: {str(device_name)[:20]:<20} 数量:{quantity} 单价:{unit_price} 公式:{formula}")
        
        # 检查合计公式
        for row in range(ws.max_row - 5, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and '合计' in str(cell_value):
                sum_formula = ws.cell(row=row, column=8).value
                print(f"\n[INFO] 合计公式 (第{row}行): {sum_formula}")
                break
        
        print(f"\n[SUCCESS] 公式验证完成!")
        
    except Exception as e:
        print(f"[ERROR] 验证失败: {str(e)}")

def main():
    """主函数"""
    print("="*60)
    print("修复Excel公式计算工具")
    print("="*60)
    
    # 修复公式
    result = fix_formula_calculation()
    
    if result:
        # 验证公式
        verify_formulas()
        
        print(f"\n[SUCCESS] 修复完成!")
        print(f"最终文件: {result['file_path']}")
        print("\n修复内容:")
        print("- 所有合价使用公式计算（=G*F格式）")
        print("- 总计使用SUM函数自动汇总")
        print("- 保持数据的动态性和可追溯性")
        print("- 修改单价或数量时，合价和总计自动更新")
        
        print(f"\n优势:")
        print("- 透明性：可以看到每个金额的计算过程")
        print("- 灵活性：修改单价或数量时自动重算")
        print("- 准确性：避免手工计算错误")
        print("- 审核性：便于财务审核和验证")
    else:
        print("[FAILED] 修复失败，请检查文件")

if __name__ == "__main__":
    main()