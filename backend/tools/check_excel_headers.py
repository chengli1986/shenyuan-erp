"""
检查Excel文件的实际表头结构
"""

import openpyxl
import os

def check_excel_headers():
    """检查Excel文件的表头"""
    
    excel_file = os.path.join(os.path.dirname(__file__), 'templates', '合同清单-完整版含两个系统-表头修正版.xlsx')
    
    if not os.path.exists(excel_file):
        print(f"[ERROR] Excel文件不存在: {excel_file}")
        return
    
    try:
        # 加载Excel文件
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        
        print("="*80)
        print("Excel表头结构检查")
        print("="*80)
        print(f"文件: {os.path.basename(excel_file)}")
        print(f"工作表: {workbook.sheetnames}")
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            print(f"\n工作表: {sheet_name}")
            
            # 检查前几行，找到表头
            for row_num in range(1, 6):
                print(f"  第{row_num}行:")
                row_values = []
                for col_num in range(1, 10):  # 检查前9列
                    cell_value = worksheet.cell(row=row_num, column=col_num).value
                    if cell_value:
                        row_values.append(str(cell_value).strip())
                    else:
                        row_values.append('')
                
                print(f"    {row_values}")
                
                # 检查是否是表头行
                if any(keyword in ' '.join(row_values) for keyword in ['序号', '设备名称', '品牌', '型号']):
                    print(f"    ↑ 这是表头行")
                    
                    # 详细分析每一列
                    print(f"    列详情:")
                    for i, value in enumerate(row_values, 1):
                        if value:
                            print(f"      第{i}列: '{value}'")
            
            # 检查第一行数据
            print(f"\n  第一行数据示例:")
            data_row = 4  # 假设数据从第4行开始
            data_values = []
            for col_num in range(1, 10):
                cell_value = worksheet.cell(row=data_row, column=col_num).value
                data_values.append(str(cell_value) if cell_value else '')
            
            for i, value in enumerate(data_values, 1):
                if value and value != '':
                    print(f"    第{i}列: '{value[:30]}...' " if len(value) > 30 else f"    第{i}列: '{value}'")
        
        workbook.close()
        
    except Exception as e:
        print(f"[ERROR] 检查失败: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    check_excel_headers()