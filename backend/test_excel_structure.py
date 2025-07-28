# 测试Excel文件结构
import openpyxl
import pandas as pd

def analyze_excel_file():
    """分析体育局投标清单Excel文件的结构"""
    file_path = "../docs/体育局投标清单.xlsx"
    
    try:
        # 使用openpyxl加载文件
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet_names = workbook.sheetnames
        
        print(f"Excel文件包含 {len(sheet_names)} 个工作表:")
        for i, sheet_name in enumerate(sheet_names, 1):
            print(f"  {i}. {sheet_name}")
        
        # 分析每个工作表
        for sheet_name in sheet_names:
            print(f"\n=== 分析工作表: {sheet_name} ===")
            worksheet = workbook[sheet_name]
            
            print(f"最大行数: {worksheet.max_row}")
            print(f"最大列数: {worksheet.max_column}")
            
            # 显示前5行数据来了解结构
            print("前5行数据:")
            for row_num in range(1, min(6, worksheet.max_row + 1)):
                row_data = []
                for col in range(1, min(worksheet.max_column + 1, 11)):  # 只显示前10列
                    cell_value = worksheet.cell(row=row_num, column=col).value
                    row_data.append(str(cell_value) if cell_value is not None else "")
                
                print(f"  第{row_num}行: {row_data}")
        
        workbook.close()
        
    except Exception as e:
        print(f"分析Excel文件失败: {str(e)}")

if __name__ == "__main__":
    analyze_excel_file()