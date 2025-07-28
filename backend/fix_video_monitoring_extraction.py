"""
修复视频监控设备提取的问题
1. 修复数据错位问题
2. 确保提取到63行的所有视频监控设备
3. 改进数据映射逻辑
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import re

def analyze_raw_excel_structure():
    """直接分析原始Excel文件的结构"""
    file_path = os.path.join(os.path.dirname(__file__), '..', 'docs', '体育局投标清单.xlsx')
    
    if not os.path.exists(file_path):
        print(f"[ERROR] 文件不存在: {file_path}")
        return None
    
    try:
        # 使用openpyxl直接读取，保持原始格式
        wb = load_workbook(file_path, data_only=True)
        ws = wb['Sheet1']  # 假设数据在第一个sheet
        
        print(f"[INFO] 直接分析Excel结构...")
        print(f"  - 最大行数: {ws.max_row}")
        print(f"  - 最大列数: {ws.max_column}")
        
        # 查找表头行
        header_row = None
        headers = []
        
        for row in range(1, min(10, ws.max_row + 1)):
            row_values = []
            for col in range(1, min(15, ws.max_column + 1)):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    row_values.append(str(cell_value).strip())
            
            # 检查是否包含表头关键词
            header_keywords = ['序号', '设备名称', '型号', '规格', '单位', '数量', '单价', '合价']
            if any(keyword in ' '.join(row_values) for keyword in header_keywords):
                header_row = row
                headers = row_values
                break
        
        if not header_row:
            print("[WARNING] 未找到明确的表头行，使用前几行数据分析")
            # 显示前10行内容用于分析
            for row in range(1, min(11, ws.max_row + 1)):
                row_data = []
                for col in range(1, min(10, ws.max_column + 1)):
                    cell_value = ws.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value else '')
                print(f"  第{row}行: {row_data}")
            
            # 手动设置表头行
            header_row = 1
            headers = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=header_row, column=col).value
                if cell_value:
                    headers.append(str(cell_value).strip())
        
        print(f"[INFO] 表头行: {header_row}")
        print(f"[INFO] 表头内容: {headers}")
        
        # 提取前65行的所有数据（确保包含到63行的视频监控设备）
        all_data = []
        video_keywords = [
            '摄像机', '摄像头', '监控', '录像机', 'NVR', 'DVR', 
            '硬盘录像', '网络录像', '视频', '监控主机',
            '球机', '枪机', '半球', '云台', '镜头',
            '视频服务器', '编码器', '解码器', '监控软件'
        ]
        
        for row in range(header_row + 1, min(66, ws.max_row + 1)):  # 提取到65行
            row_data = []
            is_video_related = False
            
            for col in range(1, len(headers) + 1):
                cell_value = ws.cell(row=row, column=col).value
                row_data.append(cell_value)
            
            # 检查是否是视频监控相关设备
            row_text = ' '.join([str(val) for val in row_data if val]).lower()
            for keyword in video_keywords:
                if keyword.lower() in row_text:
                    is_video_related = True
                    break
            
            if is_video_related:
                all_data.append({
                    'row_number': row,
                    'data': row_data,
                    'is_video': True
                })
                print(f"  第{row}行 [视频监控]: {row_data[:3]}...")  # 只显示前3列
            
            # 也保存非视频相关的数据用于调试
            if row <= 65:  # 保存前65行用于分析
                all_data.append({
                    'row_number': row,
                    'data': row_data,
                    'is_video': is_video_related
                })
        
        return {
            'headers': headers,
            'header_row': header_row,
            'all_data': all_data,
            'workbook': wb,
            'worksheet': ws
        }
        
    except Exception as e:
        print(f"[ERROR] 分析失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def map_data_to_standard_format(raw_data, headers):
    """将原始数据映射到标准格式"""
    
    # 定义列映射规则
    column_mapping = {
        '序号': ['序号', '编号', 'NO', 'No'],
        '设备名称': ['设备名称', '名称', '品名', '设备', '材料名称'],
        '品牌型号': ['型号', '品牌型号', '规格型号', '牌号', '品牌', '型号规格'],
        '技术规格参数': ['规格', '技术规格', '参数', '规格参数', '技术参数', '说明', '描述'],
        '单位': ['单位', '计量单位'],
        '数量': ['数量', '用量', '数目'],
        '综合单价': ['单价', '综合单价', '价格', '单位价格'],
        '合价': ['合价', '总价', '金额', '小计'],
        '备注': ['备注', '说明', '注释', '其他']
    }
    
    # 建立索引映射
    field_indices = {}
    for standard_field, possible_names in column_mapping.items():
        field_indices[standard_field] = None
        for i, header in enumerate(headers):
            for possible_name in possible_names:
                if possible_name in header:
                    field_indices[standard_field] = i
                    break
            if field_indices[standard_field] is not None:
                break
    
    print(f"[INFO] 字段映射结果:")
    for field, index in field_indices.items():
        header_name = headers[index] if index is not None and index < len(headers) else "未找到"
        print(f"  {field}: 第{index}列 ({header_name})")
    
    # 转换数据
    standard_data = []
    for item in raw_data:
        if not item['is_video']:
            continue
            
        data_row = item['data']
        row_num = item['row_number']
        
        # 安全获取数据
        def safe_get(index):
            if index is not None and index < len(data_row):
                return data_row[index]
            return None
        
        # 数据清理函数
        def clean_number(value):
            if value is None:
                return 0
            if isinstance(value, (int, float)):
                return float(value)
            
            # 字符串清理
            str_val = str(value).strip()
            if not str_val or str_val.lower() in ['', 'none', 'null', '-']:
                return 0
            
            # 移除非数字字符，保留小数点
            import re
            clean_val = re.sub(r'[^\d.]', '', str_val)
            try:
                return float(clean_val) if clean_val else 0
            except:
                return 0
        
        def clean_text(value):
            if value is None:
                return ''
            return str(value).strip()
        
        # 映射到标准格式
        standard_item = {
            'row_number': row_num,
            'serial_number': clean_text(safe_get(field_indices['序号'])),
            'item_name': clean_text(safe_get(field_indices['设备名称'])),
            'brand_model': clean_text(safe_get(field_indices['品牌型号'])),
            'specification': clean_text(safe_get(field_indices['技术规格参数'])),
            'unit': clean_text(safe_get(field_indices['单位'])) or '台',
            'quantity': clean_number(safe_get(field_indices['数量'])),
            'unit_price': clean_number(safe_get(field_indices['综合单价'])),
            'total_price': clean_number(safe_get(field_indices['合价'])),
            'remarks': clean_text(safe_get(field_indices['备注']))
        }
        
        # 验证数据质量
        if standard_item['item_name']:
            standard_data.append(standard_item)
            print(f"  映射第{row_num}行: {standard_item['item_name'][:20]}... 数量:{standard_item['quantity']} 单价:{standard_item['unit_price']}")
        else:
            print(f"  跳过第{row_num}行: 设备名称为空")
    
    return standard_data

def create_fixed_video_monitoring_excel(video_devices):
    """创建修复后的视频监控设备Excel文件"""
    if not video_devices:
        print("[ERROR] 没有视频监控设备数据")
        return None
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "视频监控"
    
    # v2.2模板的标准列头
    headers = ['序号', '设备名称', '品牌型号', '技术规格参数', '单位', '数量', '综合单价', '合价', '备注']
    
    # 样式定义
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    title_font = Font(bold=True, size=14)
    title_alignment = Alignment(horizontal="center", vertical="center")
    
    data_alignment = Alignment(horizontal="center", vertical="center")
    text_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 第1行：标题
    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    ws['A1'] = '体育局项目 - 视频监控系统设备清单（修复版）'
    ws['A1'].font = title_font
    ws['A1'].alignment = title_alignment
    ws.row_dimensions[1].height = 25
    
    # 第2行：说明
    ws.merge_cells(f'A2:{get_column_letter(len(headers))}2')
    ws['A2'] = f'数据来源：体育局投标清单.xlsx（1-63行），提取时间：{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws['A2'].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20
    
    # 第3行：表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
        # 设置列宽
        col_letter = get_column_letter(col_idx)
        if header == '技术规格参数':
            ws.column_dimensions[col_letter].width = 45
        elif header == '设备名称':
            ws.column_dimensions[col_letter].width = 25
        elif header == '品牌型号':
            ws.column_dimensions[col_letter].width = 20
        elif header == '备注':
            ws.column_dimensions[col_letter].width = 15
        elif header in ['序号', '单位']:
            ws.column_dimensions[col_letter].width = 8
        elif header == '数量':
            ws.column_dimensions[col_letter].width = 10
        elif header in ['综合单价', '合价']:
            ws.column_dimensions[col_letter].width = 15
        else:
            ws.column_dimensions[col_letter].width = 12
    
    ws.row_dimensions[3].height = 30
    
    # 填入设备数据
    total_amount = 0
    valid_devices = 0
    
    for row_idx, device in enumerate(video_devices, 4):
        # 确保数据正确映射
        serial_number = device.get('serial_number', '') or (row_idx - 3)
        item_name = device.get('item_name', '').strip()
        brand_model = device.get('brand_model', '').strip()
        specification = device.get('specification', '').strip()
        unit = device.get('unit', '台').strip()
        quantity = device.get('quantity', 0)
        unit_price = device.get('unit_price', 0)
        remarks = device.get('remarks', '').strip()
        
        # 跳过无效数据
        if not item_name:
            continue
        
        valid_devices += 1
        
        # 计算合价（如果原始数据有合价且合理，使用原始值）
        original_total = device.get('total_price', 0)
        if original_total > 0 and abs(original_total - (quantity * unit_price)) < 0.01:
            total_price = original_total
        else:
            total_price = quantity * unit_price if quantity and unit_price else 0
        
        total_amount += total_price
        
        # 数据行 - 确保正确的列顺序
        row_data = [
            serial_number,                    # A列：序号
            item_name,                        # B列：设备名称
            brand_model,                      # C列：品牌型号
            specification,                    # D列：技术规格参数
            unit,                            # E列：单位
            quantity,                        # F列：数量
            unit_price,                      # G列：综合单价
            f"=G{row_idx}*F{row_idx}",       # H列：合价公式
            remarks                          # I列：备注
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            # 设置对齐方式
            if col_idx in [1, 5, 6, 7, 8]:  # 序号、单位、数量、单价、合价
                cell.alignment = data_alignment
            else:
                cell.alignment = text_alignment
        
        # 添加原始行号信息到备注中（用于调试）
        if device.get('row_number'):
            original_remarks = device.get('remarks', '')
            debug_info = f"(原第{device['row_number']}行)"
            final_remarks = f"{original_remarks} {debug_info}".strip()
            ws.cell(row=row_idx, column=9, value=final_remarks)
    
    # 添加合计行
    total_row = valid_devices + 4
    ws.merge_cells(f'A{total_row}:F{total_row}')
    ws[f'A{total_row}'] = '合计'
    ws[f'A{total_row}'].font = Font(bold=True, size=12)
    ws[f'A{total_row}'].alignment = Alignment(horizontal="right", vertical="center")
    ws[f'A{total_row}'].border = thin_border
    
    # 合计金额
    ws[f'G{total_row}'] = ''
    ws[f'H{total_row}'] = f"=SUM(H4:H{total_row-1})"
    ws[f'I{total_row}'] = ''
    
    for col in ['G', 'H', 'I']:
        ws[f'{col}{total_row}'].border = thin_border
        ws[f'{col}{total_row}'].font = Font(bold=True, size=12)
        ws[f'{col}{total_row}'].alignment = data_alignment
    
    # 添加统计信息
    stats_row = total_row + 2
    ws.merge_cells(f'A{stats_row}:{get_column_letter(len(headers))}{stats_row}')
    ws[f'A{stats_row}'] = f'统计信息：共 {valid_devices} 个有效设备，预算总额：{total_amount:,.2f} 元'
    ws[f'A{stats_row}'].font = Font(size=10, color="666666")
    ws[f'A{stats_row}'].alignment = Alignment(horizontal="left", vertical="center")
    
    # 保存文件
    output_dir = os.path.join(os.path.dirname(__file__), 'templates')
    output_file = os.path.join(output_dir, '合同清单-视频监控完整内容-修复版.xlsx')
    wb.save(output_file)
    
    print(f"[OK] 修复版视频监控设备Excel文件创建成功: {output_file}")
    print(f"[INFO] 包含 {valid_devices} 个有效设备，总金额：{total_amount:,.2f} 元")
    
    return output_file

def main():
    """主函数"""
    print("="*60)
    print("体育局投标清单 - 视频监控设备提取工具（修复版）")
    print("="*60)
    
    # 1. 直接分析原始Excel结构
    print("\n第一步：直接分析原始Excel文件结构")
    raw_analysis = analyze_raw_excel_structure()
    
    if not raw_analysis:
        print("[ERROR] 原始文件分析失败，无法继续")
        return
    
    # 2. 映射数据到标准格式
    print("\n第二步：映射数据到标准格式")
    video_devices = map_data_to_standard_format(raw_analysis['all_data'], raw_analysis['headers'])
    
    if not video_devices:
        print("[ERROR] 没有找到视频监控设备")
        return
    
    print(f"\n[INFO] 成功映射 {len(video_devices)} 个视频监控设备")
    
    # 3. 生成修复版Excel文件
    print("\n第三步：生成修复版Excel文件")
    output_file = create_fixed_video_monitoring_excel(video_devices)
    
    if output_file:
        print(f"\n[SUCCESS] 修复版文件生成完成!")
        print(f"输出文件: {output_file}")
        print("\n修复内容:")
        print("- 修复了数据错位问题")
        print("- 确保提取到第63行的所有视频监控设备")
        print("- 改进了字段映射逻辑")
        print("- 添加了原始行号信息用于验证")
        
        print("\n关于AI模型识别方案:")
        print("可以考虑以下技术方案：")
        print("1. 使用OCR + NLP进行表格结构识别")
        print("2. 训练小型的表格理解模型")
        print("3. 基于规则 + 机器学习的混合方案")
        print("4. 使用现有的文档AI服务（如Azure Form Recognizer）")

if __name__ == "__main__":
    main()