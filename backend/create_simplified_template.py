"""
创建简化的Excel模板
只保留视频监控系统的完整内容，其他系统只保留表头
修复公式引用问题
"""

import pandas as pd
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def analyze_video_monitoring_sheet():
    """分析视频监控系统的详细格式和数据"""
    template_path = os.path.join(os.path.dirname(__file__), 'templates', '合同清单模板.xlsx')
    
    if not os.path.exists(template_path):
        print(f"[ERROR] 原模板文件不存在: {template_path}")
        return None
    
    try:
        wb = load_workbook(template_path)
        if '视频监控' not in wb.sheetnames:
            print("[ERROR] 视频监控工作表不存在")
            return None
        
        ws = wb['视频监控']
        
        print(f"[INFO] 分析视频监控工作表...")
        print(f"  - 最大行数: {ws.max_row}")
        print(f"  - 最大列数: {ws.max_column}")
        
        # 查找表头行
        header_row = None
        for row in range(1, min(10, ws.max_row + 1)):
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and ('序号' in str(cell_value) or '设备名称' in str(cell_value)):
                    header_row = row
                    break
            if header_row:
                break
        
        if not header_row:
            print("[ERROR] 未找到表头行")
            return None
        
        # 提取表头
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col).value
            if cell_value and str(cell_value).strip():
                headers.append(str(cell_value).strip())
            else:
                break
        
        # 移除原产地列（如果存在）
        if '原产地' in headers:
            headers.remove('原产地')
        
        print(f"[INFO] 表头 ({len(headers)}列): {headers}")
        
        # 提取设备数据
        devices = []
        for row in range(header_row + 1, ws.max_row + 1):
            row_data = []
            has_content = False
            
            for col in range(1, len(headers) + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is not None:
                    has_content = True
                row_data.append(cell_value)
            
            if has_content:
                # 跳过原产地列的数据（如果存在）
                if len(row_data) > len(headers):
                    # 假设原产地在倒数第二列
                    row_data = row_data[:len(headers)]
                devices.append(row_data)
        
        print(f"[INFO] 找到 {len(devices)} 行设备数据")
        
        return {
            'headers': headers,
            'devices': devices,
            'header_row': header_row
        }
        
    except Exception as e:
        print(f"[ERROR] 分析失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def create_simplified_template():
    """创建简化模板"""
    
    # 分析视频监控系统
    video_data = analyze_video_monitoring_sheet()
    
    if not video_data:
        print("[ERROR] 无法分析视频监控数据，使用默认格式")
        headers = ['序号', '设备名称', '品牌型号', '技术规格参数', '单位', '数量', '综合单价', '合价', '备注']
        devices = []
    else:
        headers = video_data['headers']
        devices = video_data['devices']
    
    print(f"[INFO] 使用表头: {headers}")
    print(f"[INFO] 设备数据: {len(devices)} 条")
    
    # 创建新工作簿
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认sheet
    
    # 定义系统列表
    systems = [
        ('视频监控', '视频监控系统设备清单', True),  # 保留完整数据
        ('门禁系统', '门禁对讲系统设备清单', False),  # 只保留表头
        ('综合布线', '综合布线系统材料清单', False),
        ('停车管理', '停车管理系统设备清单', False),
        ('公共广播', '公共广播系统设备清单', False),
        ('会议系统', '会议扩声系统设备清单', False),
        ('信息发布', '信息发布系统设备清单', False),
        ('楼宇自控', '楼宇自控系统设备清单', False),
        ('机房工程', '机房工程设备清单', False),
        ('UPS电源', 'UPS不间断电源设备清单', False)
    ]
    
    # 样式定义
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    title_font = Font(bold=True, size=14)
    title_alignment = Alignment(horizontal="center", vertical="center")
    
    data_alignment = Alignment(horizontal="center", vertical="center")
    text_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 找到数量、单价、合价的列索引
    quantity_col = None
    unit_price_col = None
    total_price_col = None
    
    for i, header in enumerate(headers):
        if '数量' in header:
            quantity_col = i + 1
        elif '单价' in header:
            unit_price_col = i + 1
        elif '合价' in header or '总价' in header:
            total_price_col = i + 1
    
    print(f"[INFO] 列索引 - 数量: {quantity_col}, 单价: {unit_price_col}, 合价: {total_price_col}")
    
    # 为每个系统创建工作表
    for sheet_name, sheet_title, keep_data in systems:
        ws = wb.create_sheet(title=sheet_name)
        
        # 第1行：标题
        ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
        ws['A1'] = sheet_title
        ws['A1'].font = title_font
        ws['A1'].alignment = title_alignment
        ws.row_dimensions[1].height = 25
        
        # 第2行：说明
        ws.merge_cells(f'A2:{get_column_letter(len(headers))}2')
        ws['A2'] = '说明：请按照标准格式填写设备清单，红色*标记为必填项目'
        ws['A2'].font = Font(italic=True, size=10, color="FF0000")
        ws['A2'].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 20
        
        # 第3行：表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
            # 设置列宽
            col_letter = get_column_letter(col_idx)
            if '技术规格' in header or '规格参数' in header:
                ws.column_dimensions[col_letter].width = 45
            elif '设备名称' in header:
                ws.column_dimensions[col_letter].width = 20
            elif '品牌型号' in header:
                ws.column_dimensions[col_letter].width = 18
            elif '备注' in header:
                ws.column_dimensions[col_letter].width = 15
            elif header in ['序号', '单位']:
                ws.column_dimensions[col_letter].width = 8
            elif '数量' in header:
                ws.column_dimensions[col_letter].width = 10
            elif '单价' in header or '合价' in header:
                ws.column_dimensions[col_letter].width = 12
            else:
                ws.column_dimensions[col_letter].width = 12
        
        ws.row_dimensions[3].height = 30
        
        # 如果是视频监控系统且有数据，填入数据
        if keep_data and devices:
            for row_idx, device_data in enumerate(devices, 4):
                for col_idx, value in enumerate(device_data, 1):
                    if col_idx <= len(headers):
                        # 处理合价公式
                        if col_idx == total_price_col and quantity_col and unit_price_col:
                            # 设置正确的公式：单价 × 数量
                            formula = f"={get_column_letter(unit_price_col)}{row_idx}*{get_column_letter(quantity_col)}{row_idx}"
                            cell = ws.cell(row=row_idx, column=col_idx, value=formula)
                            print(f"[INFO] 设置公式: 行{row_idx}, 列{col_idx} = {formula}")
                        else:
                            cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        
                        cell.border = thin_border
                        
                        # 设置对齐方式
                        if col_idx in [1] or (col_idx <= len(headers) and ('序号' in headers[col_idx-1] or '数量' in headers[col_idx-1] or '单价' in headers[col_idx-1] or '合价' in headers[col_idx-1])):
                            cell.alignment = data_alignment
                        else:
                            cell.alignment = text_alignment
        
        # 添加数据验证
        add_data_validation(ws, headers)
        
        # 添加页脚说明
        footer_row = 10
        ws.merge_cells(f'A{footer_row}:{get_column_letter(len(headers))}{footer_row}')
        if keep_data:
            ws[f'A{footer_row}'] = '说明：视频监控系统包含完整示例数据，其他系统请根据实际情况填写'
        else:
            ws[f'A{footer_row}'] = '说明：请在此工作表中填写对应系统的设备清单'
        ws[f'A{footer_row}'].font = Font(size=9, color="808080")
        ws[f'A{footer_row}'].alignment = Alignment(horizontal="left", vertical="center")
    
    # 保存文件
    template_dir = os.path.join(os.path.dirname(__file__), 'templates')
    file_path = os.path.join(template_dir, '合同清单简化模板_v2.2.xlsx')
    wb.save(file_path)
    
    print(f"[OK] 简化Excel模板创建成功: {file_path}")
    return file_path

def add_data_validation(ws, headers):
    """添加数据验证"""
    try:
        # 为单位列添加下拉选择
        unit_col = None
        for idx, header in enumerate(headers):
            if '单位' in header:
                unit_col = get_column_letter(idx + 1)
                break
        
        if unit_col:
            unit_validation = DataValidation(
                type="list",
                formula1='"台,个,套,对,块,根,米,箱,卷,张,片,只"',
                showErrorMessage=True,
                errorTitle="输入错误",
                error="请从下拉列表中选择单位"
            )
            ws.add_data_validation(unit_validation)
            unit_validation.add(f'{unit_col}4:{unit_col}100')
            
    except Exception as e:
        print(f"[WARNING] 添加数据验证失败: {str(e)}")

def update_guide_v22():
    """更新填写指南 v2.2"""
    guide_content = """# 合同清单Excel简化模板填写指南 v2.2

## 1. 模板特点
- **视频监控系统**：包含3个完整的示例设备，展示标准填写格式
- **其他系统**：只包含表头，需要根据实际情况填写
- **公式修正**：合价 = 单价 × 数量（公式引用已修正）

## 2. 标准列结构（9列）

| 序号 | 列名 | 是否必填 | 说明 |
|------|------|----------|------|
| A | 序号 | 建议填写 | 从1开始的数字序号 |
| B | 设备名称 | **必填** | 设备或材料的具体名称 |
| C | 品牌型号 | 建议填写 | 品牌和具体型号信息 |
| D | 技术规格参数 | 建议填写 | 详细的技术参数和规格说明 |
| E | 单位 | **必填** | 计量单位（台、个、套、米等） |
| F | 数量 | **必填** | 数量，必须是数字 |
| G | 综合单价 | **必填** | 单价（元），必须是数字 |
| H | 合价 | 建议公式 | 总价，公式：=G4*F4（单价×数量） |
| I | 备注 | 选填 | 其他说明信息 |

## 3. 公式使用规范

### 3.1 合价公式
正确的公式格式：`=G4*F4`（单价×数量）
- G列：综合单价
- F列：数量  
- H列：合价

### 3.2 示例公式
```
第4行设备：=G4*F4
第5行设备：=G5*F5
第6行设备：=G6*F6
```

## 4. 视频监控系统示例
模板中的视频监控系统包含3个设备示例：
1. 网络摄像机
2. 网络硬盘录像机  
3. 千兆交换机

这些示例展示了标准的填写格式，其他系统可参考此格式。

## 5. 填写步骤

### 5.1 选择对应系统工作表
根据设备类型选择对应的工作表：
- 视频监控、门禁系统、综合布线等

### 5.2 删除示例数据（除视频监控外）
- 保留表头行（第3行）
- 删除第4行及以后的示例数据

### 5.3 填写实际设备
- 从第4行开始填写实际设备
- 确保必填项完整
- 使用公式计算合价

## 6. 注意事项

### 6.1 公式引用
- ✅ 使用 =G4*F4 格式
- ❌ 不要使用 =F4*G4（数量×单价顺序错误）
- ✅ 每行的公式要对应正确的行号

### 6.2 数据格式
- 数量和单价必须是纯数字
- 不要在数字后加单位或货币符号
- 设备名称要具体明确

## 7. 模板版本
- 当前版本：v2.2
- 更新日期：2025-07-27
- 主要变更：
  - 修正公式引用错误（单价×数量）
  - 视频监控保留完整示例
  - 其他系统只保留表头
"""
    
    guide_path = os.path.join(os.path.dirname(__file__), 'templates', '填写指南_v2.2.md')
    with open(guide_path, 'w', encoding='utf-8') as f:
        f.write(guide_content)
    
    print(f"[OK] 填写指南v2.2创建成功: {guide_path}")

if __name__ == "__main__":
    print("开始创建简化Excel模板...")
    template_path = create_simplified_template()
    update_guide_v22()
    print("\n[SUCCESS] 简化模板创建完成!")
    print(f"新模板文件: {template_path}")
    print("特点：")
    print("- 视频监控系统：包含完整示例数据")
    print("- 其他系统：只保留表头，便于填写")
    print("- 公式修正：合价 = 单价 × 数量")