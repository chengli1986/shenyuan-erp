"""
创建更新的合同清单Excel模板
基于用户修改的视频监控格式
"""

import pandas as pd
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def analyze_video_sheet_format():
    """分析用户修改的视频监控sheet格式"""
    template_path = os.path.join(os.path.dirname(__file__), 'templates', '合同清单模板.xlsx')
    
    if not os.path.exists(template_path):
        print(f"[ERROR] 原模板文件不存在: {template_path}")
        return None
    
    try:
        wb = load_workbook(template_path)
        if '视频监控' not in wb.sheetnames:
            print("[ERROR] 视频监控工作表不存在")
            return None
        
        ws = wb['视频监控']
        
        # 分析格式
        format_info = {
            'max_row': ws.max_row,
            'max_col': ws.max_column,
            'headers': [],
            'sample_data': []
        }
        
        # 查找表头行
        header_row = None
        for row in range(1, min(10, ws.max_row + 1)):
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and '序号' in str(cell_value):
                    header_row = row
                    break
            if header_row:
                break
        
        if header_row:
            # 提取表头
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=header_row, column=col).value
                if cell_value:
                    format_info['headers'].append(str(cell_value).strip())
                else:
                    break
            
            # 提取样例数据（前3行）
            for row in range(header_row + 1, min(header_row + 4, ws.max_row + 1)):
                row_data = []
                for col in range(1, len(format_info['headers']) + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    row_data.append(cell_value)
                if any(val for val in row_data):  # 如果行不为空
                    format_info['sample_data'].append(row_data)
        
        print(f"[INFO] 分析视频监控格式完成:")
        print(f"  - 表头行: {header_row}")
        print(f"  - 列数: {len(format_info['headers'])}")
        print(f"  - 表头: {format_info['headers']}")
        print(f"  - 数据行数: {len(format_info['sample_data'])}")
        
        return format_info
        
    except Exception as e:
        print(f"[ERROR] 分析失败: {str(e)}")
        return None

def create_updated_template():
    """创建更新的模板"""
    
    # 首先分析现有格式
    video_format = analyze_video_sheet_format()
    
    if not video_format:
        print("[ERROR] 无法分析视频监控格式，使用默认格式")
        headers = ['序号', '设备名称', '品牌型号', '技术规格参数', '单位', '数量', '综合单价', '合价', '备注']
    else:
        headers = video_format['headers']
        # 如果headers中包含原产地，则移除它
        if '原产地' in headers:
            headers.remove('原产地')
            print("[INFO] 已从表头中移除原产地列")
    
    print(f"[INFO] 使用表头: {headers}")
    
    # 创建新的工作簿
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认sheet
    
    # 定义系统列表（基于真实项目文件的分析）
    systems = [
        ('视频监控', '视频监控系统设备清单'),
        ('门禁系统', '门禁对讲系统设备清单'),
        ('综合布线', '综合布线系统材料清单'),
        ('停车管理', '停车管理系统设备清单'),
        ('公共广播', '公共广播系统设备清单'),
        ('会议系统', '会议扩声系统设备清单'),
        ('信息发布', '信息发布系统设备清单'),
        ('楼宇自控', '楼宇自控系统设备清单'),
        ('机房工程', '机房工程设备清单'),
        ('UPS电源', 'UPS不间断电源设备清单')
    ]
    
    # 样式定义
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    title_font = Font(bold=True, size=14)
    title_alignment = Alignment(horizontal="center", vertical="center")
    
    data_alignment = Alignment(horizontal="center", vertical="center")
    text_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 为每个系统创建工作表
    for sheet_name, sheet_title in systems:
        ws = wb.create_sheet(title=sheet_name)
        
        # 第1行：标题
        ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
        ws['A1'] = sheet_title
        ws['A1'].font = title_font
        ws['A1'].alignment = title_alignment
        ws.row_dimensions[1].height = 25
        
        # 第2行：说明
        ws.merge_cells(f'A2:{get_column_letter(len(headers))}2')
        ws['A2'] = '说明：请按照标准格式填写设备清单，红色*标记为必填项目'
        ws['A2'].font = Font(italic=True, size=10, color="FF0000")
        ws['A2'].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 20
        
        # 第3行：表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
            # 设置列宽
            col_letter = get_column_letter(col_idx)
            if '技术规格' in header or '规格参数' in header:
                ws.column_dimensions[col_letter].width = 45
            elif '设备名称' in header:
                ws.column_dimensions[col_letter].width = 20
            elif '品牌型号' in header:
                ws.column_dimensions[col_letter].width = 18
            elif '备注' in header:
                ws.column_dimensions[col_letter].width = 15
            elif header in ['序号', '单位', '数量']:
                ws.column_dimensions[col_letter].width = 8
            elif '单价' in header or '合价' in header:
                ws.column_dimensions[col_letter].width = 12
            else:
                ws.column_dimensions[col_letter].width = 12
        
        ws.row_dimensions[3].height = 30
        
        # 如果是视频监控系统，复制用户的数据（如果有的话）
        if sheet_name == '视频监控' and video_format and video_format['sample_data']:
            for row_idx, row_data in enumerate(video_format['sample_data'], 4):
                for col_idx, value in enumerate(row_data, 1):
                    if col_idx <= len(headers):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = thin_border
                        
                        # 设置对齐方式
                        if col_idx in [1] or '序号' in headers[col_idx-1]:  # 序号
                            cell.alignment = data_alignment
                        elif col_idx in range(len(headers)) and ('数量' in headers[col_idx-1] or '单价' in headers[col_idx-1] or '合价' in headers[col_idx-1]):
                            cell.alignment = data_alignment
                        else:
                            cell.alignment = text_alignment
        else:
            # 添加示例数据行
            sample_data = get_sample_data_for_system(sheet_name, headers)
            for row_idx, row_data in enumerate(sample_data, 4):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    
                    # 设置对齐方式
                    if col_idx == 1 or (col_idx <= len(headers) and ('序号' in headers[col_idx-1] or '数量' in headers[col_idx-1] or '单价' in headers[col_idx-1] or '合价' in headers[col_idx-1])):
                        cell.alignment = data_alignment
                    else:
                        cell.alignment = text_alignment
        
        # 设置数据验证
        add_data_validation(ws, headers)
        
        # 添加页脚说明
        footer_row = 10
        ws.merge_cells(f'A{footer_row}:{get_column_letter(len(headers))}{footer_row}')
        ws[f'A{footer_row}'] = '填写提示：1.设备名称、数量、单价为必填项 2.合价建议使用公式自动计算 3.每个系统使用独立的工作表'
        ws[f'A{footer_row}'].font = Font(size=9, color="808080")
        ws[f'A{footer_row}'].alignment = Alignment(horizontal="left", vertical="center")
    
    # 保存文件
    template_dir = os.path.join(os.path.dirname(__file__), 'templates')
    file_path = os.path.join(template_dir, '合同清单标准模板_v2.1.xlsx')
    wb.save(file_path)
    
    print(f"[OK] 更新的Excel模板创建成功: {file_path}")
    return file_path

def get_sample_data_for_system(system_name, headers):
    """为不同系统生成示例数据"""
    
    # 创建基础数据结构
    base_data = {col: "" for col in headers}
    
    samples = {
        '门禁系统': [
            {**base_data, **{'序号': 1, '设备名称': '门禁控制器', '品牌型号': '熵基 ZK100', '技术规格参数': '4门控制器，TCP/IP通讯，支持刷卡+密码', '单位': '台', '数量': 3, '综合单价': 2800, '合价': '=G4*H4', '备注': '主要出入口'}},
            {**base_data, **{'序号': 2, '设备名称': '读卡器', '品牌型号': '熵基 FR100', '技术规格参数': 'IC卡读卡器，支持Mifare卡，防水设计', '单位': '台', '数量': 12, '综合单价': 350, '合价': '=G5*H5', '备注': '室外防水型'}}
        ],
        '综合布线': [
            {**base_data, **{'序号': 1, '设备名称': '六类网线', '品牌型号': '普天 CAT6', '技术规格参数': '非屏蔽双绞线，4对线芯，符合TIA/EIA标准', '单位': '箱', '数量': 20, '综合单价': 800, '合价': '=G4*H4', '备注': '305米/箱'}},
            {**base_data, **{'序号': 2, '设备名称': '信息面板', '品牌型号': '施耐德 双口', '技术规格参数': '86型面板，含2个RJ45模块', '单位': '个', '数量': 50, '综合单价': 25, '合价': '=G5*H5', '备注': '墙面安装'}}
        ],
        '停车管理': [
            {**base_data, **{'序号': 1, '设备名称': '车牌识别一体机', '品牌型号': '海康威视 DS-TCG206', '技术规格参数': '200万像素，支持车牌识别，内置补光灯', '单位': '台', '数量': 4, '综合单价': 3200, '合价': '=G4*H4', '备注': '出入口安装'}},
            {**base_data, **{'序号': 2, '设备名称': '道闸', '品牌型号': '科拓 KT-Z300', '技术规格参数': '直杆道闸，3米杆长，变频控制', '单位': '台', '数量': 4, '综合单价': 2800, '合价': '=G5*H5', '备注': '含遥控器'}}
        ]
    }
    
    # 如果有对应系统的示例数据，返回；否则返回通用示例
    if system_name in samples:
        return [list(item.values()) for item in samples[system_name]]
    else:
        # 通用示例
        return [
            [1, '示例设备1', '示例品牌型号', '请填写详细技术参数', '台', 1, 1000, '=G4*H4', '示例备注'],
            [2, '示例设备2', '示例品牌型号', '请填写详细技术参数', '台', 1, 1000, '=G5*H5', '示例备注']
        ]

def add_data_validation(ws, headers):
    """添加数据验证"""
    try:
        # 为单位列添加下拉选择
        unit_col = None
        for idx, header in enumerate(headers):
            if '单位' in header:
                unit_col = get_column_letter(idx + 1)
                break
        
        if unit_col:
            # 创建单位选项的数据验证
            unit_validation = DataValidation(
                type="list",
                formula1='"台,个,套,对,块,根,米,箱,卷,张,片,只"',
                showErrorMessage=True,
                errorTitle="输入错误",
                error="请从下拉列表中选择单位"
            )
            ws.add_data_validation(unit_validation)
            unit_validation.add(f'{unit_col}4:{unit_col}100')  # 应用到数据行
            
    except Exception as e:
        print(f"[WARNING] 添加数据验证失败: {str(e)}")

def update_guide():
    """更新填写指南"""
    guide_content = """# 合同清单Excel标准模板填写指南 v2.1

## 1. 模板结构
- **多系统工作表**：每个系统使用独立的工作表（Sheet）
- **标准化格式**：所有工作表使用统一的列结构和样式
- **自动计算**：支持Excel公式自动计算合价

## 2. 标准列结构
根据视频监控系统的格式，所有工作表包含以下列：

| 列名 | 是否必填 | 说明 |
|------|----------|------|
| 序号 | 建议填写 | 从1开始的数字序号 |
| 设备名称 | **必填** | 设备或材料的具体名称 |
| 品牌型号 | 建议填写 | 品牌和具体型号信息 |
| 技术规格参数 | 建议填写 | 详细的技术参数和规格说明 |
| 单位 | **必填** | 计量单位（台、个、套、米等） |
| 数量 | **必填** | 数量，必须是数字 |
| 综合单价 | **必填** | 单价（元），必须是数字 |
| 合价 | 建议公式 | 总价，建议使用公式 =数量×单价 |
| 备注 | 选填 | 其他说明信息 |

## 3. 填写规范

### 3.1 必填字段验证
- **设备名称**：不能为空，应具体明确
- **数量**：必须是正数，不能包含文字
- **综合单价**：必须是正数，单位为元

### 3.2 数据格式要求
```
正确示例：
设备名称：网络摄像机
数量：20
单价：1200.50

错误示例：
设备名称：（空白）
数量：20台    ❌ 不要在数字后加单位
单价：1200元  ❌ 不要在数字后加货币单位
```

### 3.3 公式使用
- **合价列**：建议使用公式 `=G4*H4`（数量×单价）
- **小计行**：可以使用 `=SUM(I4:I10)` 等汇总公式

## 4. 系统分类指南

### 4.1 预设系统模板
模板已包含以下常用系统：
- 视频监控
- 门禁系统  
- 综合布线
- 停车管理
- 公共广播
- 会议系统
- 信息发布
- 楼宇自控
- 机房工程
- UPS电源

### 4.2 自定义系统
如需添加其他系统：
1. 复制现有工作表
2. 重命名为新系统名称
3. 修改标题行
4. 删除示例数据，填入实际设备

## 5. 注意事项

### 5.1 格式要求
- ✅ 保持表头格式不变
- ✅ 使用提供的单位下拉选择
- ✅ 保持数字列的数字格式
- ❌ 不要合并单元格（标题行除外）
- ❌ 不要修改列名和列顺序

### 5.2 数据质量
- 确保设备名称具体明确
- 技术参数尽量详细完整
- 品牌型号信息准确
- 单价信息真实有效

## 6. 常见问题解答

**Q: 某些信息暂时没有怎么办？**
A: 可以留空，但必填项（设备名称、数量、单价）必须填写

**Q: 如何处理成套设备？**
A: 可以按套计算，在备注中说明包含的子设备

**Q: 单价包含安装费吗？**
A: 建议填写"综合单价"，包含设备、安装、调试等全部费用

**Q: 合价能手动填写吗？**
A: 可以，但建议使用公式自动计算，避免计算错误

**Q: 为什么删除了原产地列？**
A: 原产地信息在合同清单中使用频率低，删除后简化了表格结构

## 7. 模板版本
- 当前版本：v2.1
- 更新日期：2025-07-27
- 主要变更：删除原产地列，优化表格结构
- 基于：用户修改的视频监控系统格式
"""
    
    guide_path = os.path.join(os.path.dirname(__file__), 'templates', '填写指南_v2.md')
    with open(guide_path, 'w', encoding='utf-8') as f:
        f.write(guide_content)
    
    print(f"[OK] 更新的填写指南创建成功: {guide_path}")

if __name__ == "__main__":
    print("开始创建更新的Excel模板...")
    template_path = create_updated_template()
    update_guide()
    print("\n[SUCCESS] 模板更新完成!")
    print(f"新模板文件: {template_path}")
    print("建议用户使用新模板进行数据录入，以确保最佳的解析效果。")