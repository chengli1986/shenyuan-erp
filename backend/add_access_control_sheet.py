"""
在Excel文件中添加第二个工作表：出入口控制
自动生成标准表头格式
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def add_access_control_sheet():
    """添加出入口控制工作表"""
    
    # 使用最终修复版文件
    file_path = os.path.join(os.path.dirname(__file__), 'templates', '合同清单-视频监控完整内容-最终修复版.xlsx')
    
    if not os.path.exists(file_path):
        print(f"[ERROR] 文件不存在: {file_path}")
        return None
    
    try:
        # 加载工作簿
        wb = load_workbook(file_path)
        
        print(f"[INFO] 添加出入口控制工作表...")
        print(f"[INFO] 基础文件: {os.path.basename(file_path)}")
        print(f"[INFO] 当前工作表: {wb.sheetnames}")
        
        # 创建新的工作表
        ws_access = wb.create_sheet(title="出入口控制")
        
        # 标准列头（与视频监控保持一致）
        headers = ['序号', '设备名称', '品牌型号', '技术规格参数', '单位', '数量', '综合单价', '合价', '备注']
        
        # 样式定义
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        title_font = Font(bold=True, size=14)
        title_alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 第1行：标题
        ws_access.merge_cells(f'A1:{get_column_letter(len(headers))}1')
        ws_access['A1'] = '体育局项目 - 出入口控制系统设备清单'
        ws_access['A1'].font = title_font
        ws_access['A1'].alignment = title_alignment
        ws_access.row_dimensions[1].height = 25
        
        # 第2行：说明
        ws_access.merge_cells(f'A2:{get_column_letter(len(headers))}2')
        ws_access['A2'] = '说明：请按照标准格式填写出入口控制系统设备清单，包括门禁、考勤、停车等相关设备'
        ws_access['A2'].font = Font(italic=True, size=10, color="FF0000")
        ws_access['A2'].alignment = Alignment(horizontal="left", vertical="center")
        ws_access.row_dimensions[2].height = 20
        
        # 第3行：表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws_access.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
            # 设置列宽
            col_letter = get_column_letter(col_idx)
            if header == '技术规格参数':
                ws_access.column_dimensions[col_letter].width = 45
            elif header == '设备名称':
                ws_access.column_dimensions[col_letter].width = 25
            elif header == '品牌型号':
                ws_access.column_dimensions[col_letter].width = 20
            elif header == '备注':
                ws_access.column_dimensions[col_letter].width = 15
            elif header in ['序号', '单位']:
                ws_access.column_dimensions[col_letter].width = 8
            elif header == '数量':
                ws_access.column_dimensions[col_letter].width = 10
            elif header in ['综合单价', '合价']:
                ws_access.column_dimensions[col_letter].width = 15
            else:
                ws_access.column_dimensions[col_letter].width = 12
        
        ws_access.row_dimensions[3].height = 30
        
        # 添加示例数据行（空行模板）
        example_data = [
            ['1', '门禁控制器', '熵基 ZK100', '4门控制器，TCP/IP通讯，支持刷卡+密码', '台', '3', '2800', '=G4*F4', '主要出入口'],
            ['2', '读卡器', '熵基 FR100', 'IC卡读卡器，支持Mifare卡，防水设计', '台', '12', '350', '=G5*F5', '室外防水型'],
            ['3', '电锁', '阿萨盟 8600', '电磁锁，12V/24V，锁力280kg', '把', '8', '280', '=G6*F6', '双门锁'],
            ['4', 'IC卡', 'Mifare 1K', '13.56MHz，可重复擦写1万次', '张', '200', '5', '=G7*F7', '员工卡'],
            ['5', '门禁电源', '12V/3A', '开关电源，带后备电池功能', '台', '6', '180', '=G8*F8', '带蓄电池']
        ]
        
        # 写入示例数据
        for row_idx, row_data in enumerate(example_data, 4):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_access.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                
                # 设置对齐方式
                if col_idx in [1, 5, 6, 7, 8]:  # 序号、单位、数量、单价、合价
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                # 设置数字格式
                if col_idx == 6:  # 数量
                    cell.number_format = '0.00'
                elif col_idx in [7, 8]:  # 单价、合价
                    cell.number_format = '#,##0.00'
        
        # 添加合计行
        total_row = 4 + len(example_data) + 1
        
        total_font = Font(bold=True, size=12)
        total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        total_alignment = Alignment(horizontal="center", vertical="center")
        amount_alignment = Alignment(horizontal="right", vertical="center")
        
        # 合并单元格并添加"合计"
        ws_access.merge_cells(f'A{total_row}:F{total_row}')
        total_cell = ws_access[f'A{total_row}']
        total_cell.value = '合计'
        total_cell.font = total_font
        total_cell.fill = total_fill
        total_cell.alignment = total_alignment
        total_cell.border = thin_border
        
        # 添加合计公式
        sum_formula = f"=SUM(H4:H{total_row-1})"
        amount_cell = ws_access[f'H{total_row}']
        amount_cell.value = sum_formula
        amount_cell.font = total_font
        amount_cell.fill = total_fill
        amount_cell.alignment = amount_alignment
        amount_cell.border = thin_border
        amount_cell.number_format = '#,##0.00'
        
        # G列和I列样式
        for col in ['G', 'I']:
            cell = ws_access[f'{col}{total_row}']
            cell.border = thin_border
            cell.fill = total_fill
        
        # 添加说明信息
        info_row = total_row + 2
        ws_access.merge_cells(f'A{info_row}:I{info_row}')
        info_cell = ws_access[f'A{info_row}']
        info_cell.value = '填写说明：上述为示例数据，请根据实际项目需求删除示例并填写真实的出入口控制设备信息'
        info_cell.font = Font(size=10, italic=True, color="666666")
        info_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # 设备类别提示
        category_row = info_row + 1
        ws_access.merge_cells(f'A{category_row}:I{category_row}')
        category_cell = ws_access[f'A{category_row}']
        category_cell.value = '设备类别：门禁控制器、读卡器、电锁、门磁、按钮、IC卡、考勤机、停车道闸、车牌识别等'
        category_cell.font = Font(size=10, color="333333")
        category_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        print(f"[INFO] 出入口控制工作表创建完成")
        print(f"[INFO] 添加了 {len(example_data)} 行示例数据")
        print(f"[INFO] 合计公式: {sum_formula}")
        
        # 保存文件
        output_path = os.path.join(os.path.dirname(__file__), 'templates', '合同清单-完整版含两个系统.xlsx')
        wb.save(output_path)
        
        print(f"\n[OK] 双系统Excel文件创建成功!")
        print(f"[INFO] 输出文件: {output_path}")
        print(f"[INFO] 工作表数量: {len(wb.sheetnames)}")
        print(f"[INFO] 工作表列表: {wb.sheetnames}")
        
        return {
            'file_path': output_path,
            'worksheets': wb.sheetnames,
            'example_rows': len(example_data)
        }
        
    except Exception as e:
        print(f"[ERROR] 添加工作表失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def verify_dual_system_file():
    """验证双系统文件"""
    file_path = os.path.join(os.path.dirname(__file__), 'templates', '合同清单-完整版含两个系统.xlsx')
    
    if not os.path.exists(file_path):
        print("[ERROR] 双系统文件不存在")
        return
    
    try:
        wb = load_workbook(file_path)
        
        print(f"\n[INFO] 验证双系统文件:")
        print(f"  - 文件名: {os.path.basename(file_path)}")
        print(f"  - 工作表数: {len(wb.sheetnames)}")
        
        # 验证每个工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n[INFO] 工作表: {sheet_name}")
            
            # 检查标题
            title = ws['A1'].value
            print(f"  - 标题: {title}")
            
            # 检查表头
            headers = []
            for col in range(1, 10):
                header = ws.cell(row=3, column=col).value
                if header:
                    headers.append(header)
            print(f"  - 表头: {', '.join(headers)}")
            
            # 检查数据行数
            data_rows = 0
            for row in range(4, ws.max_row + 1):
                device_name = ws.cell(row=row, column=2).value
                if device_name and '合计' not in str(device_name):
                    data_rows += 1
                elif '合计' in str(device_name):
                    break
            print(f"  - 数据行数: {data_rows}")
            
            # 检查合计公式
            for row in range(ws.max_row - 5, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and '合计' in str(cell_value):
                    formula = ws.cell(row=row, column=8).value
                    print(f"  - 合计公式: {formula}")
                    break
        
        print(f"\n[SUCCESS] 双系统文件验证完成!")
        
    except Exception as e:
        print(f"[ERROR] 验证失败: {str(e)}")

def main():
    """主函数"""
    print("="*80)
    print("添加出入口控制系统工作表")
    print("="*80)
    
    # 添加新工作表
    result = add_access_control_sheet()
    
    if result:
        # 验证结果
        verify_dual_system_file()
        
        print(f"\n[SUCCESS] 出入口控制工作表添加完成!")
        print(f"输出文件: {result['file_path']}")
        print(f"\n文件特点:")
        print(f"- 包含 {len(result['worksheets'])} 个工作表: {', '.join(result['worksheets'])}")
        print(f"- 出入口控制包含 {result['example_rows']} 行示例数据")
        print(f"- 标准9列格式，与视频监控保持一致")
        print(f"- 自动计算公式和合计功能")
        print(f"- 包含设备类别提示和填写说明")
        
        print(f"\n使用说明:")
        print(f"1. 视频监控工作表：包含49项真实设备数据")
        print(f"2. 出入口控制工作表：包含示例数据，可删除后填入实际数据")
        print(f"3. 两个工作表格式完全一致，便于统一管理")
        print(f"4. 可以继续添加更多系统工作表")
    else:
        print("[FAILED] 添加工作表失败")

if __name__ == "__main__":
    main()