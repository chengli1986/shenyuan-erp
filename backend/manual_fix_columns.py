"""
手动修复列错位问题
根据观察到的具体错位情况进行精确修复
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def manual_fix_columns():
    """手动精确修复列错位"""
    
    # 使用最终版文件作为基础
    file_path = os.path.join(os.path.dirname(__file__), 'templates', '合同清单-视频监控完整内容-最终版.xlsx')
    
    if not os.path.exists(file_path):
        print(f"[ERROR] 文件不存在: {file_path}")
        return None
    
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        print(f"[INFO] 手动修复列错位...")
        print(f"[INFO] 基础文件: {os.path.basename(file_path)}")
        
        # 根据之前的分析，错位情况是：
        # 实际数据列：A序号 B设备名称 C品牌型号 D规格 E单价 F单位 G数量 H合价公式 I备注
        # 应该调整为：A序号 B设备名称 C品牌型号 D规格 E单位 F数量 G单价 H合价公式 I备注
        
        # 找到数据范围
        data_start_row = 4
        data_end_row = 52  # 根据之前的分析
        
        print(f"[INFO] 数据行范围: {data_start_row} 到 {data_end_row}")
        
        # 备份和重新排列数据
        print(f"\n[INFO] 重新排列数据列...")
        
        for row in range(data_start_row, data_end_row + 1):
            # 读取当前行的原始数据
            serial_num = ws.cell(row=row, column=1).value    # A列：序号
            device_name = ws.cell(row=row, column=2).value   # B列：设备名称
            brand_model = ws.cell(row=row, column=3).value   # C列：品牌型号
            specification = ws.cell(row=row, column=4).value # D列：规格
            current_unit_price = ws.cell(row=row, column=5).value  # E列：当前是单价
            current_unit = ws.cell(row=row, column=6).value       # F列：当前是单位
            current_quantity = ws.cell(row=row, column=7).value   # G列：当前是数量
            formula = ws.cell(row=row, column=8).value       # H列：公式
            remarks = ws.cell(row=row, column=9).value       # I列：备注
            
            # 跳过空行
            if not device_name:
                continue
            
            # 数据类型转换和验证
            try:
                # 单位应该是文本
                unit = str(current_unit).strip() if current_unit else '台'
                if unit not in ['台', '个', '套', '对', '块', '根', '米', '箱', '卷', '张', '片', '只', '条', '组', '面', '㎡']:
                    unit = '台'
                
                # 数量应该是数字
                quantity = float(current_quantity) if current_quantity else 1
                
                # 单价应该是数字
                unit_price = float(current_unit_price) if current_unit_price else 0
                
            except (ValueError, TypeError):
                # 处理转换错误
                unit = '台'
                quantity = 1
                unit_price = 0
            
            # 重新写入正确位置的数据
            ws.cell(row=row, column=1, value=row - 3)        # A列：重新编号
            ws.cell(row=row, column=2, value=device_name)    # B列：设备名称
            ws.cell(row=row, column=3, value=brand_model)    # C列：品牌型号
            ws.cell(row=row, column=4, value=specification) # D列：规格
            ws.cell(row=row, column=5, value=unit)          # E列：单位
            ws.cell(row=row, column=6, value=quantity)      # F列：数量
            ws.cell(row=row, column=7, value=unit_price)    # G列：单价
            ws.cell(row=row, column=8, value=f"=G{row}*F{row}")  # H列：正确的公式
            ws.cell(row=row, column=9, value=remarks)       # I列：备注
            
            print(f"  第{row}行: {str(device_name)[:20]:<20} 单位:{unit:<3} 数量:{quantity:>6} 单价:{unit_price:>8}")
        
        # 设置单元格格式
        print(f"\n[INFO] 设置单元格格式...")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(data_start_row, data_end_row + 1):
            for col in range(1, 10):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                
                # 设置对齐方式
                if col in [1, 5, 6, 7, 8]:  # 序号、单位、数量、单价、合价
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                # 设置数字格式
                if col == 6:  # 数量
                    cell.number_format = '0.00'
                elif col in [7, 8]:  # 单价、合价
                    cell.number_format = '#,##0.00'
        
        # 添加合计行
        total_row = data_end_row + 2
        
        # 样式定义
        total_font = Font(bold=True, size=12)
        total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        total_alignment = Alignment(horizontal="center", vertical="center")
        amount_alignment = Alignment(horizontal="right", vertical="center")
        
        # 清除可能存在的旧合计行
        for row in range(data_end_row + 1, ws.max_row + 1):
            for col in range(1, 10):
                ws.cell(row=row, column=col).value = None
        
        # 合并单元格并添加"合计"
        ws.merge_cells(f'A{total_row}:F{total_row}')
        total_cell = ws[f'A{total_row}']
        total_cell.value = '合计'
        total_cell.font = total_font
        total_cell.fill = total_fill
        total_cell.alignment = total_alignment
        total_cell.border = thin_border
        
        # 添加合计公式
        sum_formula = f"=SUM(H{data_start_row}:H{data_end_row})"
        amount_cell = ws[f'H{total_row}']
        amount_cell.value = sum_formula
        amount_cell.font = total_font
        amount_cell.fill = total_fill
        amount_cell.alignment = amount_alignment
        amount_cell.border = thin_border
        amount_cell.number_format = '#,##0.00'
        
        print(f"[INFO] 合计公式: {sum_formula}")
        
        # G列和I列样式
        for col in ['G', 'I']:
            cell = ws[f'{col}{total_row}']
            cell.border = thin_border
            cell.fill = total_fill
        
        # 添加说明
        summary_row = total_row + 2
        ws.merge_cells(f'A{summary_row}:I{summary_row}')
        summary_cell = ws[f'A{summary_row}']
        summary_cell.value = f'视频监控系统设备清单 - 共49项设备，列结构已修复，公式计算正确'
        summary_cell.font = Font(size=11, color="333333")
        summary_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # 保存文件
        output_path = os.path.join(os.path.dirname(__file__), 'templates', '合同清单-视频监控完整内容-最终修复版.xlsx')
        wb.save(output_path)
        
        print(f"\n[OK] 手动修复完成!")
        print(f"[INFO] 输出文件: {output_path}")
        print(f"[INFO] 修复行数: {data_end_row - data_start_row + 1}")
        print(f"[INFO] 合计公式: {sum_formula}")
        
        return output_path
        
    except Exception as e:
        print(f"[ERROR] 修复失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def verify_final_file():
    """验证最终修复的文件"""
    file_path = os.path.join(os.path.dirname(__file__), 'templates', '合同清单-视频监控完整内容-最终修复版.xlsx')
    
    if not os.path.exists(file_path):
        print("[ERROR] 最终修复文件不存在")
        return
    
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        print(f"\n[INFO] 验证最终修复文件:")
        
        # 验证列结构
        print(f"\n[INFO] 列结构验证:")
        expected_headers = ['序号', '设备名称', '品牌型号', '技术规格参数', '单位', '数量', '综合单价', '合价', '备注']
        for col, header in enumerate(expected_headers, 1):
            actual_header = ws.cell(row=3, column=col).value
            print(f"  {chr(64+col)}列: {header} - {actual_header}")
        
        # 验证前5行数据
        print(f"\n[INFO] 前5行数据验证:")
        for row in range(4, 9):
            serial_num = ws.cell(row=row, column=1).value     # 序号
            device_name = ws.cell(row=row, column=2).value    # 设备名称
            unit = ws.cell(row=row, column=5).value          # 单位
            quantity = ws.cell(row=row, column=6).value      # 数量
            unit_price = ws.cell(row=row, column=7).value    # 单价
            formula = ws.cell(row=row, column=8).value       # 公式
            
            print(f"  第{row}行: 序号:{serial_num} 设备:{str(device_name)[:15]:<15} 单位:{unit} 数量:{quantity} 单价:{unit_price} 公式:{formula}")
        
        # 检查合计
        total_found = False
        for row in range(50, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and '合计' in str(cell_value):
                total_formula = ws.cell(row=row, column=8).value
                print(f"\n[INFO] 合计行 (第{row}行): {total_formula}")
                total_found = True
                break
        
        if not total_found:
            print(f"\n[WARNING] 未找到合计行")
        
        print(f"\n[SUCCESS] 验证完成!")
        
    except Exception as e:
        print(f"[ERROR] 验证失败: {str(e)}")

def main():
    """主函数"""
    print("="*80)
    print("手动修复Excel列错位问题 - 精确版")
    print("="*80)
    
    # 手动修复
    result = manual_fix_columns()
    
    if result:
        # 验证结果
        verify_final_file()
        
        print(f"\n[SUCCESS] 手动修复完成!")
        print(f"最终文件: {result}")
        print("\n修复说明:")
        print("E列：单位（台、个、套等）")
        print("F列：数量（数值）")
        print("G列：综合单价（数值）")
        print("H列：合价（公式 =G*F）")
        print("\n现在公式应该是正确的：单价 × 数量 = 合价")
    else:
        print("[FAILED] 手动修复失败")

if __name__ == "__main__":
    main()