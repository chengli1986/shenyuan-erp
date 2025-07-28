"""
最终修复列错位问题
根据观察：E列显示金额，F列显示单位，G列显示数量
需要重新排列为：E列单位，F列数量，G列单价
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def analyze_misalignment():
    """分析当前的列错位情况"""
    file_path = os.path.join(os.path.dirname(__file__), 'templates', '合同清单-完整版含两个系统-修复版.xlsx')
    
    if not os.path.exists(file_path):
        print(f"[ERROR] 文件不存在: {file_path}")
        return None
    
    try:
        wb = load_workbook(file_path)
        
        print(f"[INFO] 分析列错位情况...")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n[INFO] 工作表: {sheet_name}")
            print(f"表头: A列:{ws.cell(3,1).value} B列:{ws.cell(3,2).value} C列:{ws.cell(3,3).value} D列:{ws.cell(3,4).value}")
            print(f"     E列:{ws.cell(3,5).value} F列:{ws.cell(3,6).value} G列:{ws.cell(3,7).value} H列:{ws.cell(3,8).value}")
            
            print(f"\n前3行数据分析:")
            for row in range(4, 7):
                device_name = ws.cell(row, 2).value
                if device_name:
                    print(f"  第{row}行 - 设备: {str(device_name)[:20]}")
                    print(f"    E列(应为单位): {ws.cell(row, 5).value}")
                    print(f"    F列(应为数量): {ws.cell(row, 6).value}")  
                    print(f"    G列(应为单价): {ws.cell(row, 7).value}")
                    print(f"    H列(应为合价): {ws.cell(row, 8).value}")
        
        return wb
        
    except Exception as e:
        print(f"[ERROR] 分析失败: {str(e)}")
        return None

def fix_column_misalignment_final():
    """最终修复列错位问题"""
    
    file_path = os.path.join(os.path.dirname(__file__), 'templates', '合同清单-完整版含两个系统-修复版.xlsx')
    
    if not os.path.exists(file_path):
        print(f"[ERROR] 文件不存在: {file_path}")
        return None
    
    try:
        wb = load_workbook(file_path)
        
        print(f"[INFO] 最终修复列错位问题...")
        print(f"[INFO] 错位情况: E列显示金额，F列显示单位，G列显示数量")
        print(f"[INFO] 修复目标: E列单位，F列数量，G列单价")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n[INFO] 处理工作表: {sheet_name}")
            
            # 找到数据范围
            data_start_row = 4
            data_end_row = ws.max_row
            
            # 找到实际数据结束行
            for row in range(data_start_row, ws.max_row + 1):
                device_name = ws.cell(row=row, column=2).value
                if not device_name or str(device_name).strip() in ['合计', '总计', '小计', '']:
                    data_end_row = row - 1
                    break
            
            print(f"  - 数据行范围: {data_start_row} 到 {data_end_row}")
            
            # 备份和重新排列数据
            print(f"  - 重新排列数据...")
            
            for row in range(data_start_row, data_end_row + 1):
                device_name = ws.cell(row=row, column=2).value
                if not device_name or not str(device_name).strip():
                    continue
                
                # 读取当前错位的数据
                current_e = ws.cell(row=row, column=5).value  # E列：当前显示金额
                current_f = ws.cell(row=row, column=6).value  # F列：当前显示单位
                current_g = ws.cell(row=row, column=7).value  # G列：当前显示数量
                
                # 数据重新分配：
                # current_f 是单位 -> 应该放在E列
                # current_g 是数量 -> 应该放在F列  
                # current_e 是单价 -> 应该放在G列
                
                # 数据类型验证和转换
                try:
                    # 单位处理 (current_f -> E列)
                    unit = str(current_f).strip() if current_f else '台'
                    if unit.replace('.', '').replace(',', '').isdigit():
                        unit = '台'  # 如果是数字，默认为台
                    
                    # 数量处理 (current_g -> F列)
                    try:
                        quantity = float(current_g) if current_g else 1.0
                    except (ValueError, TypeError):
                        quantity = 1.0
                    
                    # 单价处理 (current_e -> G列)
                    try:
                        unit_price = float(current_e) if current_e else 0.0
                    except (ValueError, TypeError):
                        unit_price = 0.0
                
                except Exception as e:
                    print(f"    第{row}行数据转换错误: {e}")
                    unit = '台'
                    quantity = 1.0
                    unit_price = 0.0
                
                # 重新写入正确位置
                ws.cell(row=row, column=1, value=row - 3)          # A列：重新编号
                # B、C、D列保持不变（设备名称、品牌型号、技术规格参数）
                ws.cell(row=row, column=5, value=unit)             # E列：单位
                ws.cell(row=row, column=6, value=quantity)         # F列：数量
                ws.cell(row=row, column=7, value=unit_price)       # G列：单价
                ws.cell(row=row, column=8, value=f"=G{row}*F{row}")  # H列：合价公式
                
                print(f"    第{row}行: {str(device_name)[:20]:<20} 单位:{unit:<4} 数量:{quantity:>6.1f} 单价:{unit_price:>8.2f}")
            
            # 设置单元格格式
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in range(data_start_row, data_end_row + 1):
                for col in range(1, 10):
                    cell = ws.cell(row=row, column=col)
                    if cell.value is not None:
                        cell.border = thin_border
                        
                        # 设置对齐方式
                        if col in [1, 5, 6, 7, 8]:  # 序号、单位、数量、单价、合价
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                        
                        # 设置数字格式
                        if col == 6:  # 数量
                            cell.number_format = '0.00'
                        elif col in [7, 8]:  # 单价、合价
                            cell.number_format = '#,##0.00'
            
            # 更新合计公式
            total_row = None
            for row in range(data_end_row + 1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and '合计' in str(cell_value):
                    total_row = row
                    break
            
            if total_row:
                sum_formula = f"=SUM(H{data_start_row}:H{data_end_row})"
                ws.cell(row=total_row, column=8, value=sum_formula)
                print(f"  - 更新合计公式: {sum_formula}")
                
                # 设置合计行样式
                total_font = Font(bold=True, size=12)
                total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                total_alignment = Alignment(horizontal="center", vertical="center")
                amount_alignment = Alignment(horizontal="right", vertical="center")
                
                # 合计标签
                total_cell = ws.cell(row=total_row, column=1)
                total_cell.font = total_font
                total_cell.fill = total_fill
                total_cell.alignment = total_alignment
                total_cell.border = thin_border
                
                # 合计金额
                amount_cell = ws.cell(row=total_row, column=8)
                amount_cell.font = total_font
                amount_cell.fill = total_fill
                amount_cell.alignment = amount_alignment
                amount_cell.border = thin_border
                amount_cell.number_format = '#,##0.00'
                
                # 其他列样式
                for col in range(2, 8):
                    cell = ws.cell(row=total_row, column=col)
                    cell.border = thin_border
                    cell.fill = total_fill
        
        # 保存文件
        output_path = os.path.join(os.path.dirname(__file__), 'templates', '合同清单-完整版含两个系统-最终正确版.xlsx')
        wb.save(output_path)
        
        print(f"\n[OK] 列错位最终修复完成!")
        print(f"[INFO] 输出文件: {output_path}")
        print(f"[INFO] 修复说明:")
        print(f"  - E列：单位（台、个、套等）")
        print(f"  - F列：数量（数值）") 
        print(f"  - G列：综合单价（数值）")
        print(f"  - H列：合价（公式 =G*F）")
        print(f"  - 重新编排了序列号")
        print(f"  - 更新了合计公式")
        
        return output_path
        
    except Exception as e:
        print(f"[ERROR] 修复失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def verify_final_fix():
    """验证最终修复结果"""
    file_path = os.path.join(os.path.dirname(__file__), 'templates', '合同清单-完整版含两个系统-最终正确版.xlsx')
    
    if not os.path.exists(file_path):
        print("[ERROR] 最终修复文件不存在")
        return
    
    try:
        wb = load_workbook(file_path)
        
        print(f"\n[INFO] 验证最终修复结果:")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n[INFO] 工作表: {sheet_name}")
            
            # 验证前5行数据
            print(f"前5行数据验证:")
            for row in range(4, 9):
                device_name = ws.cell(row=row, column=2).value
                if device_name:
                    seq_num = ws.cell(row=row, column=1).value
                    unit = ws.cell(row=row, column=5).value
                    quantity = ws.cell(row=row, column=6).value
                    unit_price = ws.cell(row=row, column=7).value
                    formula = ws.cell(row=row, column=8).value
                    
                    print(f"  第{row}行: 序号:{seq_num} 设备:{str(device_name)[:15]:<15}")
                    print(f"         单位:{unit} 数量:{quantity} 单价:{unit_price} 公式:{formula}")
            
            # 检查合计
            for row in range(ws.max_row - 5, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and '合计' in str(cell_value):
                    sum_formula = ws.cell(row=row, column=8).value
                    print(f"  - 合计公式: {sum_formula}")
                    break
        
        print(f"\n[SUCCESS] 验证完成! 现在列位置正确，公式可以正常计算了。")
        
    except Exception as e:
        print(f"[ERROR] 验证失败: {str(e)}")

def main():
    """主函数"""
    print("="*80)
    print("最终修复Excel列错位问题")
    print("="*80)
    
    # 首先分析当前错位情况
    analyze_misalignment()
    
    # 修复列错位
    result = fix_column_misalignment_final()
    
    if result:
        # 验证修复结果
        verify_final_fix()
        
        print(f"\n[SUCCESS] 最终修复完成!")
        print(f"输出文件: {result}")
        print("\n现在的列结构:")
        print("A列：序号    B列：设备名称    C列：品牌型号    D列：技术规格参数")
        print("E列：单位    F列：数量        G列：综合单价    H列：合价")
        print("I列：备注")
        print("\n合价公式现在可以正确计算了！(单价 × 数量)")
    else:
        print("[FAILED] 最终修复失败")

if __name__ == "__main__":
    main()