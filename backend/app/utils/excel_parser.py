# backend/app/utils/excel_parser.py
"""
Excel文件解析工具类

用于解析投标清单Excel文件，支持多sheet结构
根据用户提供的Excel格式进行解析，提取设备信息
"""

import pandas as pd
import openpyxl
from typing import Dict, List, Optional, Tuple, Any
from decimal import Decimal
import logging
from pathlib import Path

# 配置日志
logger = logging.getLogger(__name__)

class ExcelParseError(Exception):
    """Excel解析异常"""
    pass

class ContractExcelParser:
    """
    合同清单Excel解析器
    
    支持解析投标清单的Excel文件，提取系统分类和设备明细信息
    """
    
    def __init__(self):
        """初始化解析器"""
        self.workbook = None
        self.sheet_names = []
        self.parsed_data = {
            'categories': [],  # 系统分类信息
            'items': [],       # 设备明细信息
            'summary': {}      # 汇总信息
        }
    
    def load_excel_file(self, file_path: str) -> bool:
        """
        加载Excel文件
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            bool: 是否加载成功
        """
        try:
            if not Path(file_path).exists():
                raise ExcelParseError(f"文件不存在: {file_path}")
            
            # 使用openpyxl加载Excel文件
            self.workbook = openpyxl.load_workbook(file_path, data_only=True)
            self.sheet_names = self.workbook.sheetnames
            
            logger.info(f"成功加载Excel文件: {file_path}")
            logger.info(f"工作表列表: {self.sheet_names}")
            
            return True
            
        except Exception as e:
            logger.error(f"加载Excel文件失败: {str(e)}")
            raise ExcelParseError(f"加载Excel文件失败: {str(e)}")
    
    def detect_header_row(self, worksheet, max_scan_rows: int = 20) -> int:
        """
        自动检测表头所在行
        
        Args:
            worksheet: 工作表对象
            max_scan_rows: 最大扫描行数
            
        Returns:
            int: 表头行号(从1开始)
        """
        # 常见的表头关键词
        header_keywords = [
            '序号', '设备名称', '品牌型号', '规格', '单位', '数量', '单价',
            '总价', '金额', '备注', '原产地', '型号', '名称'
        ]
        
        for row_num in range(1, min(max_scan_rows + 1, worksheet.max_row + 1)):
            row_values = []
            for col in range(1, min(20, worksheet.max_column + 1)):  # 检查前20列
                cell_value = worksheet.cell(row=row_num, column=col).value
                if cell_value:
                    row_values.append(str(cell_value).strip())
            
            # 检查这一行是否包含足够多的表头关键词
            keyword_count = sum(1 for keyword in header_keywords 
                              if any(keyword in value for value in row_values))
            
            if keyword_count >= 3:  # 至少包含3个关键词
                logger.info(f"检测到表头在第 {row_num} 行: {row_values}")
                return row_num
        
        # 默认返回第1行
        logger.warning("未能自动检测到表头，使用第1行作为表头")
        return 1
    
    def parse_sheet_data(self, sheet_name: str) -> Dict[str, Any]:
        """
        解析单个工作表的数据
        
        Args:
            sheet_name: 工作表名称
            
        Returns:
            dict: 解析后的数据
        """
        if not self.workbook:
            raise ExcelParseError("请先加载Excel文件")
        
        if sheet_name not in self.sheet_names:
            raise ExcelParseError(f"工作表 '{sheet_name}' 不存在")
        
        worksheet = self.workbook[sheet_name]
        
        # 检测表头行
        header_row = self.detect_header_row(worksheet)
        
        # 直接从worksheet读取数据，转换为DataFrame
        try:
            # 获取所有数据
            data = []
            headers = []
            
            # 读取表头
            for col in range(1, worksheet.max_column + 1):
                header_cell = worksheet.cell(row=header_row, column=col)
                headers.append(header_cell.value if header_cell.value else f'col_{col}')
            
            # 读取数据行
            for row_num in range(header_row + 1, worksheet.max_row + 1):
                row_data = []
                has_data = False
                
                for col in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=row_num, column=col).value
                    row_data.append(cell_value)
                    if cell_value is not None:
                        has_data = True
                
                if has_data:  # 只添加有数据的行
                    data.append(row_data)
            
            # 创建DataFrame
            df = pd.DataFrame(data, columns=headers)
            
            # 清理数据
            df = self._clean_dataframe(df)
            
            # 解析设备明细
            items = self._parse_items_from_dataframe(df, sheet_name)
            
            return {
                'sheet_name': sheet_name,
                'header_row': header_row,
                'total_rows': len(df),
                'items': items,
                'raw_data': df.to_dict('records') if len(df) < 1000 else None  # 大文件不返回原始数据
            }
            
        except Exception as e:
            logger.error(f"解析工作表 '{sheet_name}' 失败: {str(e)}")
            raise ExcelParseError(f"解析工作表 '{sheet_name}' 失败: {str(e)}")
    
    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        清理DataFrame数据
        
        Args:
            df: 原始DataFrame
            
        Returns:
            DataFrame: 清理后的数据
        """
        # 删除完全空白的行
        df = df.dropna(how='all')
        
        # 重置索引
        df = df.reset_index(drop=True)
        
        # 标准化列名
        df.columns = [self._normalize_column_name(col) for col in df.columns]
        
        return df
    
    def _normalize_column_name(self, col_name: str) -> str:
        """
        标准化列名
        
        Args:
            col_name: 原始列名
            
        Returns:
            str: 标准化后的列名
        """
        if pd.isna(col_name):
            return 'unnamed_column'
        
        col_str = str(col_name).strip()
        
        # 列名映射表
        column_mapping = {
            '序号': 'serial_number',
            '设备名称': 'item_name',
            '名称': 'item_name',
            '品牌型号': 'brand_model',
            '型号': 'brand_model',
            '规格': 'specification',
            '规格描述': 'specification',
            '单位': 'unit',
            '数量': 'quantity',
            '单价': 'unit_price',
            '价格': 'unit_price',
            '总价': 'total_price',
            '金额': 'total_price',
            '原产地': 'origin_place',
            '产地': 'origin_place',
            '备注': 'remarks',
            '说明': 'remarks',
            '物料类型': 'item_type',
            '类型': 'item_type'
        }
        
        # 查找匹配的标准列名
        for key, value in column_mapping.items():
            if key in col_str:
                return value
        
        # 如果没有匹配，返回原始列名的简化版本
        return col_str.replace(' ', '_').lower()
    
    def _parse_items_from_dataframe(self, df: pd.DataFrame, sheet_name: str) -> List[Dict[str, Any]]:
        """
        从DataFrame中解析设备明细
        
        Args:
            df: 数据DataFrame
            sheet_name: 工作表名称
            
        Returns:
            list: 设备明细列表
        """
        items = []
        
        for index, row in df.iterrows():
            try:
                # 跳过空行或无效行
                if self._is_empty_row(row):
                    continue
                
                # 构建设备明细数据
                item = self._build_item_from_row(row, sheet_name, index + 1)
                
                if item:
                    items.append(item)
                    
            except Exception as e:
                logger.warning(f"解析第 {index + 1} 行数据失败: {str(e)}")
                continue
        
        logger.info(f"从工作表 '{sheet_name}' 解析出 {len(items)} 个设备明细")
        return items
    
    def _is_empty_row(self, row: pd.Series) -> bool:
        """
        判断是否为空行
        
        Args:
            row: 数据行
            
        Returns:
            bool: 是否为空行
        """
        # 检查重要字段是否为空
        important_fields = ['item_name', 'brand_model', 'quantity']
        
        for field in important_fields:
            if field in row and pd.notna(row[field]) and str(row[field]).strip():
                return False
        
        return True
    
    def _build_item_from_row(self, row: pd.Series, sheet_name: str, row_number: int) -> Optional[Dict[str, Any]]:
        """
        从数据行构建设备明细
        
        Args:
            row: 数据行
            sheet_name: 工作表名称
            row_number: 行号
            
        Returns:
            dict: 设备明细数据
        """
        try:
            # 提取基本信息
            item = {
                'serial_number': self._safe_get_value(row, 'serial_number', str(row_number)),
                'item_name': self._safe_get_value(row, 'item_name', '').strip(),
                'brand_model': self._safe_get_value(row, 'brand_model', '').strip(),
                'specification': self._safe_get_value(row, 'specification', '').strip(),
                'unit': self._safe_get_value(row, 'unit', '台').strip(),
                'quantity': self._parse_decimal(row, 'quantity'),
                'unit_price': self._parse_decimal(row, 'unit_price'),
                'origin_place': self._safe_get_value(row, 'origin_place', '中国').strip(),
                'item_type': self._determine_item_type(row),
                'remarks': self._safe_get_value(row, 'remarks', '').strip(),
                'excel_sheet_name': sheet_name,
                'excel_row_number': row_number
            }
            
            # 验证必要字段
            if not item['item_name']:
                logger.warning(f"第 {row_number} 行缺少设备名称，跳过")
                return None
            
            if item['quantity'] is None or item['quantity'] <= 0:
                logger.warning(f"第 {row_number} 行数量无效: {item['quantity']}，跳过")
                return None
            
            # 计算总价
            if item['unit_price'] is not None:
                item['total_price'] = item['quantity'] * item['unit_price']
            else:
                # 尝试从total_price列获取
                item['total_price'] = self._parse_decimal(row, 'total_price')
            
            return item
            
        except Exception as e:
            logger.error(f"构建第 {row_number} 行设备明细失败: {str(e)}")
            return None
    
    def _safe_get_value(self, row: pd.Series, column: str, default: Any = None) -> Any:
        """
        安全获取行数据
        
        Args:
            row: 数据行
            column: 列名
            default: 默认值
            
        Returns:
            Any: 列值或默认值
        """
        if column in row and pd.notna(row[column]):
            value = row[column]
            # 将值转换为字符串并strip
            if value is not None:
                return str(value).strip()
            return value
        return default
    
    def _parse_decimal(self, row: pd.Series, column: str) -> Optional[Decimal]:
        """
        解析数字字段
        
        Args:
            row: 数据行
            column: 列名
            
        Returns:
            Decimal: 解析后的数字或None
        """
        value = self._safe_get_value(row, column)
        
        if value is None:
            return None
        
        try:
            # 清理数字字符串
            if isinstance(value, str):
                # 移除常见的非数字字符
                value = value.replace(',', '').replace('￥', '').replace('¥', '').replace('元', '').strip()
                if not value:
                    return None
            
            return Decimal(str(value))
            
        except (ValueError, TypeError, Exception):
            logger.warning(f"无法解析数字: {value}")
            return None
    
    def _determine_item_type(self, row: pd.Series) -> str:
        """
        确定物料类型
        
        Args:
            row: 数据行
            
        Returns:
            str: 物料类型（主材或辅材）
        """
        # 如果有item_type列，直接使用
        item_type = self._safe_get_value(row, 'item_type', '').strip()
        if item_type in ['主材', '辅材']:
            return item_type
        
        # 根据设备名称判断
        item_name = self._safe_get_value(row, 'item_name', '').lower()
        
        # 辅材关键词
        auxiliary_keywords = [
            '线材', '电缆', '管材', '支架', '配件', '螺丝', '工具',
            '安装', '调试', '维护', '辅助', '配套'
        ]
        
        for keyword in auxiliary_keywords:
            if keyword in item_name:
                return '辅材'
        
        # 默认为主材
        return '主材'
    
    def parse_all_sheets(self) -> Dict[str, Any]:
        """
        解析所有工作表
        
        Returns:
            dict: 解析结果
        """
        if not self.workbook:
            raise ExcelParseError("请先加载Excel文件")
        
        all_results = {
            'categories': [],
            'items': [],
            'sheets_info': [],
            'summary': {
                'total_sheets': len(self.sheet_names),
                'total_items': 0,
                'total_amount': Decimal('0'),
                'parse_errors': []
            }
        }
        
        for sheet_name in self.sheet_names:
            try:
                logger.info(f"开始解析工作表: {sheet_name}")
                
                sheet_result = self.parse_sheet_data(sheet_name)
                
                # 创建系统分类
                category = {
                    'category_name': sheet_name,
                    'category_code': self._generate_category_code(sheet_name),
                    'excel_sheet_name': sheet_name,
                    'total_items_count': len(sheet_result['items']),
                    'budget_amount': sum(
                        item.get('total_price', 0) for item in sheet_result['items']
                        if item.get('total_price')
                    ),
                    'description': f"从Excel工作表 '{sheet_name}' 导入的系统分类"
                }
                
                all_results['categories'].append(category)
                
                # 添加设备明细
                for item in sheet_result['items']:
                    item['category_name'] = sheet_name
                    item['category_code'] = category['category_code']
                
                all_results['items'].extend(sheet_result['items'])
                all_results['sheets_info'].append(sheet_result)
                
                # 更新汇总信息
                all_results['summary']['total_items'] += len(sheet_result['items'])
                
                logger.info(f"工作表 '{sheet_name}' 解析完成，设备数量: {len(sheet_result['items'])}")
                
            except Exception as e:
                error_msg = f"解析工作表 '{sheet_name}' 失败: {str(e)}"
                logger.error(error_msg)
                all_results['summary']['parse_errors'].append(error_msg)
        
        # 计算总金额
        all_results['summary']['total_amount'] = sum(
            item.get('total_price', 0) for item in all_results['items']
            if item.get('total_price')
        )
        
        logger.info(f"Excel解析完成，总计: {all_results['summary']['total_items']} 个设备，"
                   f"{all_results['summary']['total_amount']} 元")
        
        return all_results
    
    def _generate_category_code(self, sheet_name: str) -> str:
        """
        生成系统分类编码
        
        Args:
            sheet_name: 工作表名称
            
        Returns:
            str: 系统分类编码
        """
        # 简单的编码生成逻辑，可以根据需要改进
        import re
        
        # 移除特殊字符，保留中英文和数字
        clean_name = re.sub(r'[^\w\u4e00-\u9fff]', '', sheet_name)
        
        # 生成编码
        code = f"SYS_{clean_name.upper()}"
        
        return code[:50]  # 限制长度
    
    def get_parsed_data(self) -> Dict[str, Any]:
        """
        获取解析后的数据
        
        Returns:
            dict: 解析后的数据
        """
        return self.parsed_data.copy()
    
    def close(self):
        """关闭工作簿"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None

# 便捷函数
def parse_contract_excel(file_path: str) -> Dict[str, Any]:
    """
    解析合同清单Excel文件的便捷函数
    
    Args:
        file_path: Excel文件路径
        
    Returns:
        dict: 解析结果
    """
    parser = ContractExcelParser()
    try:
        parser.load_excel_file(file_path)
        return parser.parse_all_sheets()
    finally:
        parser.close()

if __name__ == "__main__":
    # 测试代码
    import sys
    import os
    
    # 添加项目根目录到路径
    sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
    
    # 测试解析
    test_file = r"C:\Users\ch_w1\shenyuan-erp\docs\体育局投标清单.xlsx"
    
    if os.path.exists(test_file):
        print(f"测试解析文件: {test_file}")
        try:
            result = parse_contract_excel(test_file)
            print(f"解析成功！")
            print(f"总工作表数: {result['summary']['total_sheets']}")
            print(f"总设备数: {result['summary']['total_items']}")
            print(f"总金额: {result['summary']['total_amount']}")
            print(f"系统分类: {[cat['category_name'] for cat in result['categories']]}")
        except Exception as e:
            print(f"测试失败: {str(e)}")
    else:
        print(f"测试文件不存在: {test_file}")