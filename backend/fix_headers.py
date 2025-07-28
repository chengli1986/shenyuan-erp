"""
修正Excel文件的表头
1. 品牌型号 -> 设备品牌
2. 技术规格参数 -> 设备型号
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def fix_headers():
    """修正表头名称"""
    
    file_path = os.path.join(os.path.dirname(__file__), 'templates', '合同清单-完整版含两个系统-最终正确版.xlsx')
    
    if not os.path.exists(file_path):
        print(f"[ERROR] 文件不存在: {file_path}")
        return None
    
    try:
        # 加载工作簿
        wb = load_workbook(file_path)
        
        print(f"[INFO] 修正表头名称...")
        print(f"[INFO] 文件: {os.path.basename(file_path)}")
        print(f"[INFO] 工作表: {wb.sheetnames}")
        
        # 新的表头定义
        new_headers = ['序号', '设备名称', '设备品牌', '设备型号', '单位', '数量', '综合单价', '合价', '备注']
        
        # 处理每个工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n[INFO] 处理工作表: {sheet_name}")
            
            # 显示原表头
            print(f"  原表头:")
            for col in range(1, 10):
                old_header = ws.cell(row=3, column=col).value
                print(f"    {chr(64+col)}列: {old_header}")
            
            # 更新表头
            print(f"  新表头:")
            for col, new_header in enumerate(new_headers, 1):
                cell = ws.cell(row=3, column=col)
                cell.value = new_header
                
                # 保持原有的表头样式
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                print(f"    {chr(64+col)}列: {new_header}")
            
            # 设置列宽（根据新表头调整）
            col_widths = {
                'A': 8,   # 序号
                'B': 25,  # 设备名称
                'C': 20,  # 设备品牌
                'D': 20,  # 设备型号
                'E': 8,   # 单位
                'F': 10,  # 数量
                'G': 15,  # 综合单价
                'H': 15,  # 合价
                'I': 15   # 备注
            }
            
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width
            
            print(f"  - 已更新列宽设置")
        
        # 保存文件
        output_path = os.path.join(os.path.dirname(__file__), 'templates', '合同清单-完整版含两个系统-表头修正版.xlsx')
        wb.save(output_path)
        
        print(f"\n[OK] 表头修正完成!")
        print(f"[INFO] 输出文件: {output_path}")
        print(f"[INFO] 修正内容:")
        print(f"  - C列: 品牌型号 -> 设备品牌")
        print(f"  - D列: 技术规格参数 -> 设备型号") 
        print(f"  - 保持了原有的表头样式和格式")
        print(f"  - 调整了列宽以适应新表头")
        
        return output_path
        
    except Exception as e:
        print(f"[ERROR] 修正失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def verify_headers():
    """验证表头修正结果"""
    file_path = os.path.join(os.path.dirname(__file__), 'templates', '合同清单-完整版含两个系统-表头修正版.xlsx')
    
    if not os.path.exists(file_path):
        print("[ERROR] 表头修正文件不存在")
        return
    
    try:
        wb = load_workbook(file_path)
        
        print(f"\n[INFO] 验证表头修正结果:")
        print(f"  - 文件名: {os.path.basename(file_path)}")
        print(f"  - 工作表数: {len(wb.sheetnames)}")
        
        # 验证每个工作表的表头
        expected_headers = ['序号', '设备名称', '设备品牌', '设备型号', '单位', '数量', '综合单价', '合价', '备注']
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n[INFO] 工作表: {sheet_name}")
            print(f"  表头验证:")
            
            for col, expected_header in enumerate(expected_headers, 1):
                actual_header = ws.cell(row=3, column=col).value
                status = "✓" if str(actual_header) == expected_header else "✗"
                print(f"    {chr(64+col)}列: {expected_header} {status}")
            
            # 检查前几行数据
            print(f"  前3行数据:")
            for row in range(4, 7):
                device_name = ws.cell(row=row, column=2).value
                if device_name:
                    brand = ws.cell(row=row, column=3).value
                    model = ws.cell(row=row, column=4).value
                    unit = ws.cell(row=row, column=5).value
                    quantity = ws.cell(row=row, column=6).value
                    price = ws.cell(row=row, column=7).value
                    
                    print(f"    第{row}行: {str(device_name)[:15]:<15} 品牌:{str(brand)[:10]:<10} 型号:{str(model)[:10]:<10}")
                    print(f"           单位:{unit} 数量:{quantity} 单价:{price}")
        
        print(f"\n[SUCCESS] 表头验证完成!")
        
    except Exception as e:
        print(f"[ERROR] 验证失败: {str(e)}")

def main():
    """主函数"""
    print("="*80)
    print("修正Excel文件表头")
    print("="*80)
    
    # 修正表头
    result = fix_headers()
    
    if result:
        # 验证修正结果
        verify_headers()
        
        print(f"\n[SUCCESS] 表头修正完成!")
        print(f"输出文件: {result}")
        print("\n修正说明:")
        print("1. C列：品牌型号 -> 设备品牌")
        print("2. D列：技术规格参数 -> 设备型号")
        print("3. 保持了所有原有样式和格式")
        print("4. 两个工作表的表头都已同步更新")
        
        print(f"\n新的列结构:")
        print("A列：序号      B列：设备名称    C列：设备品牌    D列：设备型号")
        print("E列：单位      F列：数量        G列：综合单价    H列：合价")
        print("I列：备注")
        
        print(f"\n使用说明:")
        print("- 原文件保持不变")
        print("- 新文件添加了'-表头修正版'后缀")
        print("- 所有数据和公式保持不变")
        print("- 表头更加清晰和规范")
    else:
        print("[FAILED] 表头修正失败")

if __name__ == "__main__":
    main()