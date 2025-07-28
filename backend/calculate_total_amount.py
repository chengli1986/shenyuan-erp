"""
计算视频监控设备清单的总金额
并添加汇总信息到Excel文件
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def calculate_and_add_total():
    """计算总金额并添加到Excel文件"""
    
    # 文件路径
    file_path = os.path.join(os.path.dirname(__file__), 'templates', '合同清单-视频监控完整内容-最终版.xlsx')
    
    if not os.path.exists(file_path):
        print(f"[ERROR] 文件不存在: {file_path}")
        return None
    
    try:
        # 加载工作簿
        wb = load_workbook(file_path)
        ws = wb.active
        
        print(f"[INFO] 正在计算总金额...")
        print(f"[INFO] 文件: {os.path.basename(file_path)}")
        
        # 找到数据范围
        data_start_row = 4
        data_end_row = ws.max_row
        
        # 找到实际的数据结束行
        for row in range(data_start_row, ws.max_row + 1):
            device_name = ws.cell(row=row, column=2).value
            if not device_name or str(device_name).strip() in ['合计', '总计', '小计', '']:
                data_end_row = row - 1
                break
        
        print(f"[INFO] 数据行范围: {data_start_row} 到 {data_end_row}")
        
        # 计算总金额
        total_amount = 0
        device_count = 0
        invalid_rows = []
        
        print(f"\n[INFO] 逐行计算金额:")
        for row in range(data_start_row, data_end_row + 1):
            device_name = ws.cell(row=row, column=2).value
            quantity = ws.cell(row=row, column=6).value  # F列：数量
            unit_price = ws.cell(row=row, column=7).value  # G列：单价
            total_price = ws.cell(row=row, column=8).value  # H列：合价
            
            if device_name and str(device_name).strip():
                device_count += 1
                
                # 尝试获取合价（如果是公式，计算结果）
                if total_price and isinstance(total_price, str) and total_price.startswith('='):
                    # 如果是公式，手动计算
                    try:
                        q = float(quantity) if quantity else 0
                        p = float(unit_price) if unit_price else 0
                        calculated_total = q * p
                        total_amount += calculated_total
                        print(f"  第{row}行: {str(device_name)[:25]:<25} {q:>6} × {p:>8,.2f} = {calculated_total:>10,.2f}")
                    except (ValueError, TypeError):
                        invalid_rows.append(row)
                        print(f"  第{row}行: {str(device_name)[:25]:<25} [计算错误] 数量:{quantity} 单价:{unit_price}")
                else:
                    # 如果有直接的合价值
                    try:
                        amount = float(total_price) if total_price else 0
                        total_amount += amount
                        print(f"  第{row}行: {str(device_name)[:25]:<25} 合价: {amount:>10,.2f}")
                    except (ValueError, TypeError):
                        # 尝试用数量×单价计算
                        try:
                            q = float(quantity) if quantity else 0
                            p = float(unit_price) if unit_price else 0
                            calculated_total = q * p
                            total_amount += calculated_total
                            print(f"  第{row}行: {str(device_name)[:25]:<25} {q:>6} × {p:>8,.2f} = {calculated_total:>10,.2f}")
                        except (ValueError, TypeError):
                            invalid_rows.append(row)
                            print(f"  第{row}行: {str(device_name)[:25]:<25} [数据无效] 数量:{quantity} 单价:{unit_price}")
        
        print(f"\n[INFO] 计算结果:")
        print(f"  - 有效设备数: {device_count}")
        print(f"  - 无效数据行: {len(invalid_rows)}")
        if invalid_rows:
            print(f"  - 无效行号: {invalid_rows}")
        print(f"  - 总金额: {total_amount:,.2f} 元")
        
        # 添加合计行
        total_row = data_end_row + 2
        
        # 样式定义
        total_font = Font(bold=True, size=12)
        total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        total_alignment = Alignment(horizontal="center", vertical="center")
        amount_alignment = Alignment(horizontal="right", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 合并单元格并添加"合计"
        ws.merge_cells(f'A{total_row}:F{total_row}')
        total_cell = ws[f'A{total_row}']
        total_cell.value = '合计'
        total_cell.font = total_font
        total_cell.fill = total_fill
        total_cell.alignment = total_alignment
        total_cell.border = thin_border
        
        # 添加总金额
        amount_cell = ws[f'H{total_row}']
        amount_cell.value = total_amount
        amount_cell.font = total_font
        amount_cell.fill = total_fill
        amount_cell.alignment = amount_alignment
        amount_cell.border = thin_border
        amount_cell.number_format = '#,##0.00'
        
        # G列和I列也加上边框和样式
        for col in ['G', 'I']:
            cell = ws[f'{col}{total_row}']
            cell.border = thin_border
            cell.fill = total_fill
        
        # 添加统计摘要
        summary_row = total_row + 2
        ws.merge_cells(f'A{summary_row}:I{summary_row}')
        summary_cell = ws[f'A{summary_row}']
        summary_cell.value = f'项目摘要：视频监控系统共 {device_count} 项设备，总投资预算：{total_amount:,.2f} 元（￥{total_amount:,.2f}）'
        summary_cell.font = Font(size=11, color="333333")
        summary_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # 添加大写金额
        chinese_amount = number_to_chinese(total_amount)
        chinese_row = summary_row + 1
        ws.merge_cells(f'A{chinese_row}:I{chinese_row}')
        chinese_cell = ws[f'A{chinese_row}']
        chinese_cell.value = f'大写金额：{chinese_amount}'
        chinese_cell.font = Font(size=11, bold=True, color="333333")
        chinese_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # 保存文件
        output_path = os.path.join(os.path.dirname(__file__), 'templates', '合同清单-视频监控完整内容-含总金额.xlsx')
        wb.save(output_path)
        
        print(f"\n[OK] 总金额计算完成!")
        print(f"[INFO] 输出文件: {output_path}")
        print(f"[INFO] 设备总数: {device_count} 项")
        print(f"[INFO] 总金额: ￥{total_amount:,.2f}")
        print(f"[INFO] 大写金额: {chinese_amount}")
        
        return {
            'file_path': output_path,
            'device_count': device_count,
            'total_amount': total_amount,
            'chinese_amount': chinese_amount,
            'invalid_rows': invalid_rows
        }
        
    except Exception as e:
        print(f"[ERROR] 计算失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def number_to_chinese(amount):
    """将数字转换为中文大写金额"""
    if amount == 0:
        return "零元整"
    
    # 中文数字
    chinese_digits = ['零', '壹', '贰', '叁', '肆', '伍', '陆', '柒', '捌', '玖']
    chinese_units = ['', '拾', '佰', '仟', '万', '拾', '佰', '仟', '亿']
    
    # 分离整数和小数部分
    yuan = int(amount)
    jiao = int((amount - yuan) * 10)
    fen = int(((amount - yuan) * 100) % 10)
    
    result = ""
    
    # 处理元的部分
    if yuan == 0:
        result = "零元"
    else:
        yuan_str = str(yuan)
        yuan_len = len(yuan_str)
        
        for i, digit in enumerate(yuan_str):
            digit_val = int(digit)
            unit_index = yuan_len - i - 1
            
            if digit_val != 0:
                result += chinese_digits[digit_val] + chinese_units[unit_index]
            elif result and not result.endswith('零'):
                result += chinese_digits[0]
        
        # 处理万和亿的特殊情况
        if '万' in result and result.endswith('零'):
            result = result[:-1]
        
        result += "元"
    
    # 处理角分
    if jiao == 0 and fen == 0:
        result += "整"
    else:
        if jiao != 0:
            result += chinese_digits[jiao] + "角"
        if fen != 0:
            result += chinese_digits[fen] + "分"
    
    return result

def display_summary():
    """显示金额汇总信息"""
    print("\n" + "="*80)
    print("视频监控系统设备清单 - 金额汇总报告")
    print("="*80)
    
    result = calculate_and_add_total()
    
    if result:
        print(f"\n📋 项目概览:")
        print(f"   设备数量: {result['device_count']:>10} 项")
        print(f"   总投资额: ￥{result['total_amount']:>12,.2f}")
        print(f"   大写金额: {result['chinese_amount']}")
        
        if result['invalid_rows']:
            print(f"\n⚠️  数据警告:")
            print(f"   无效数据行: {len(result['invalid_rows'])} 行")
            print(f"   行号: {result['invalid_rows']}")
        
        print(f"\n📁 输出文件:")
        print(f"   {result['file_path']}")
        
        print(f"\n✅ 处理完成!")
        
        return result
    else:
        print("\n❌ 处理失败!")
        return None

if __name__ == "__main__":
    display_summary()