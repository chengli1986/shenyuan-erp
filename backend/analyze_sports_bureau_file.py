"""
分析体育局投标清单.xlsx文件
提取视频监控相关设备，生成标准模板格式的Excel文件
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from app.utils.excel_parser import ContractExcelParser
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import re

def analyze_sports_bureau_file():
    """分析体育局投标清单文件"""
    file_path = os.path.join(os.path.dirname(__file__), '..', 'docs', '体育局投标清单.xlsx')
    
    if not os.path.exists(file_path):
        print(f"[ERROR] 文件不存在: {file_path}")
        return None
    
    print(f"[INFO] 分析文件: {file_path}")
    print(f"[INFO] 文件大小: {os.path.getsize(file_path) / 1024:.2f} KB")
    
    parser = ContractExcelParser()
    
    try:
        # 加载Excel文件
        parser.load_excel_file(file_path)
        print(f"[INFO] 工作表数量: {len(parser.sheet_names)}")
        print(f"[INFO] 工作表列表: {parser.sheet_names}")
        
        # 解析所有工作表
        result = parser.parse_all_sheets()
        
        categories = result.get('categories', [])
        items = result.get('items', [])
        summary = result.get('summary', {})
        
        print(f"\n[INFO] 解析结果:")
        print(f"  - 系统分类数: {len(categories)}")
        print(f"  - 设备总数: {len(items)}")
        print(f"  - 总金额: {summary.get('total_amount', 0):,.2f}")
        
        # 显示每个分类
        print(f"\n[INFO] 系统分类详情:")
        for i, cat in enumerate(categories, 1):
            print(f"  {i:2d}. {cat['category_name']:<15} | "
                  f"设备数: {cat['total_items_count']:>3} | "
                  f"金额: {float(cat['budget_amount']):>12,.2f}")
        
        return {
            'categories': categories,
            'items': items,
            'summary': summary
        }
        
    except Exception as e:
        print(f"[ERROR] 分析失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def extract_video_monitoring_devices(analysis_result):
    """提取视频监控相关设备"""
    if not analysis_result:
        return []
    
    items = analysis_result['items']
    video_devices = []
    
    # 视频监控相关关键词
    video_keywords = [
        '摄像机', '摄像头', '监控', '录像机', 'NVR', 'DVR', 
        '硬盘录像', '网络录像', '视频', '监控主机',
        '球机', '枪机', '半球', '云台', '镜头',
        '视频服务器', '编码器', '解码器'
    ]
    
    print(f"\n[INFO] 开始提取视频监控设备...")
    
    for item in items:
        item_name = item.get('item_name', '').strip()
        specification = item.get('specification', '').strip()
        brand_model = item.get('brand_model', '').strip()
        
        # 检查设备名称、规格参数、品牌型号中是否包含视频监控关键词
        content_to_check = f"{item_name} {specification} {brand_model}".lower()
        
        is_video_device = False
        for keyword in video_keywords:
            if keyword.lower() in content_to_check:
                is_video_device = True
                break
        
        if is_video_device:
            video_devices.append(item)
            print(f"  - 找到设备: {item_name}")
    
    print(f"\n[INFO] 共找到 {len(video_devices)} 个视频监控设备")
    
    # 按设备类型分组统计
    device_types = {}
    for device in video_devices:
        item_name = device.get('item_name', '未知设备')
        if item_name in device_types:
            device_types[item_name] += 1
        else:
            device_types[item_name] = 1
    
    print(f"\n[INFO] 设备类型统计:")
    for device_type, count in device_types.items():
        print(f"  - {device_type}: {count} 台/套")
    
    return video_devices

def create_video_monitoring_excel(video_devices):
    """创建视频监控设备的Excel文件"""
    if not video_devices:
        print("[ERROR] 没有视频监控设备数据")
        return None
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "视频监控"
    
    # v2.2模板的标准列头
    headers = ['序号', '设备名称', '品牌型号', '技术规格参数', '单位', '数量', '综合单价', '合价', '备注']
    
    # 样式定义
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    title_font = Font(bold=True, size=14)
    title_alignment = Alignment(horizontal="center", vertical="center")
    
    data_alignment = Alignment(horizontal="center", vertical="center")
    text_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 第1行：标题
    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    ws['A1'] = '体育局项目 - 视频监控系统设备清单'
    ws['A1'].font = title_font
    ws['A1'].alignment = title_alignment
    ws.row_dimensions[1].height = 25
    
    # 第2行：说明
    ws.merge_cells(f'A2:{get_column_letter(len(headers))}2')
    ws['A2'] = f'数据来源：体育局投标清单.xlsx，提取时间：{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws['A2'].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20
    
    # 第3行：表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
        # 设置列宽
        col_letter = get_column_letter(col_idx)
        if header == '技术规格参数':
            ws.column_dimensions[col_letter].width = 50
        elif header == '设备名称':
            ws.column_dimensions[col_letter].width = 25
        elif header == '品牌型号':
            ws.column_dimensions[col_letter].width = 20
        elif header == '备注':
            ws.column_dimensions[col_letter].width = 20
        elif header in ['序号', '单位']:
            ws.column_dimensions[col_letter].width = 8
        elif header == '数量':
            ws.column_dimensions[col_letter].width = 10
        elif header in ['综合单价', '合价']:
            ws.column_dimensions[col_letter].width = 15
        else:
            ws.column_dimensions[col_letter].width = 12
    
    ws.row_dimensions[3].height = 30
    
    # 填入设备数据
    total_amount = 0
    
    for row_idx, device in enumerate(video_devices, 4):
        # 提取数据
        serial_number = row_idx - 3
        item_name = device.get('item_name', '').strip()
        brand_model = device.get('brand_model', '').strip()
        specification = device.get('specification', '').strip()
        unit = device.get('unit', '台').strip()
        quantity = device.get('quantity', 1)
        unit_price = device.get('unit_price', 0)
        remarks = device.get('remarks', '').strip()
        
        # 计算合价
        if isinstance(quantity, (int, float)) and isinstance(unit_price, (int, float)):
            total_price = quantity * unit_price
        else:
            total_price = 0
        
        total_amount += total_price
        
        # 数据行
        row_data = [
            serial_number,                    # 序号
            item_name,                        # 设备名称
            brand_model,                      # 品牌型号
            specification,                    # 技术规格参数
            unit,                            # 单位
            quantity,                        # 数量
            unit_price,                      # 综合单价
            f"=G{row_idx}*F{row_idx}",       # 合价公式
            remarks                          # 备注
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            # 设置对齐方式
            if col_idx in [1, 5, 6, 7, 8]:  # 序号、数量、单价、合价
                cell.alignment = data_alignment
            else:
                cell.alignment = text_alignment
    
    # 添加合计行
    total_row = len(video_devices) + 4
    ws.merge_cells(f'A{total_row}:F{total_row}')
    ws[f'A{total_row}'] = '合计'
    ws[f'A{total_row}'].font = Font(bold=True, size=12)
    ws[f'A{total_row}'].alignment = Alignment(horizontal="right", vertical="center")
    ws[f'A{total_row}'].border = thin_border
    
    # 合计金额
    ws[f'G{total_row}'] = ''
    ws[f'H{total_row}'] = f"=SUM(H4:H{total_row-1})"
    ws[f'I{total_row}'] = ''
    
    for col in ['G', 'H', 'I']:
        ws[f'{col}{total_row}'].border = thin_border
        ws[f'{col}{total_row}'].font = Font(bold=True, size=12)
        ws[f'{col}{total_row}'].alignment = data_alignment
    
    # 添加统计信息
    stats_row = total_row + 2
    ws.merge_cells(f'A{stats_row}:{get_column_letter(len(headers))}{stats_row}')
    ws[f'A{stats_row}'] = f'统计信息：共 {len(video_devices)} 个设备，预算总额：{total_amount:,.2f} 元'
    ws[f'A{stats_row}'].font = Font(size=10, color="666666")
    ws[f'A{stats_row}'].alignment = Alignment(horizontal="left", vertical="center")
    
    # 保存文件
    output_dir = os.path.join(os.path.dirname(__file__), 'templates')
    output_file = os.path.join(output_dir, '合同清单-视频监控完整内容.xlsx')
    wb.save(output_file)
    
    print(f"[OK] 视频监控设备Excel文件创建成功: {output_file}")
    print(f"[INFO] 包含 {len(video_devices)} 个设备，总金额：{total_amount:,.2f} 元")
    
    return output_file

def main():
    """主函数"""
    print("="*60)
    print("体育局投标清单 - 视频监控设备提取工具")
    print("="*60)
    
    # 1. 分析原始文件
    print("\n第一步：分析体育局投标清单文件")
    analysis_result = analyze_sports_bureau_file()
    
    if not analysis_result:
        print("[ERROR] 文件分析失败，无法继续")
        return
    
    # 2. 提取视频监控设备
    print("\n第二步：提取视频监控相关设备")
    video_devices = extract_video_monitoring_devices(analysis_result)
    
    if not video_devices:
        print("[ERROR] 没有找到视频监控设备")
        return
    
    # 3. 生成标准格式Excel文件
    print("\n第三步：生成标准模板格式的Excel文件")
    output_file = create_video_monitoring_excel(video_devices)
    
    if output_file:
        print(f"\n[SUCCESS] 任务完成!")
        print(f"输出文件: {output_file}")
        print("\n文件特点:")
        print("- 符合v2.2标准模板格式")
        print("- 包含正确的合价公式")
        print("- 提取真实项目数据")
        print("- 可直接用于项目报价")

if __name__ == "__main__":
    main()