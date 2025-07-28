"""
修复Excel文件中的序列号，从1开始重新排序
"""

import os
from openpyxl import load_workbook

def fix_serial_numbers():
    """修复序列号，从1开始重新排序"""
    
    # 文件路径
    file_path = os.path.join(os.path.dirname(__file__), 'templates', '合同清单-视频监控完整内容-修复版.xlsx')
    
    if not os.path.exists(file_path):
        print(f"[ERROR] 文件不存在: {file_path}")
        return None
    
    try:
        # 加载工作簿
        wb = load_workbook(file_path)
        ws = wb.active
        
        print(f"[INFO] 正在修复文件: {os.path.basename(file_path)}")
        print(f"[INFO] 工作表名称: {ws.title}")
        print(f"[INFO] 最大行数: {ws.max_row}")
        
        # 找到数据开始行（假设第4行开始是数据）
        data_start_row = 4
        
        # 找到数据结束行（查找合计行或空行）
        data_end_row = ws.max_row
        for row in range(data_start_row, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value is None or str(cell_value).strip() in ['合计', '总计', '小计']:
                data_end_row = row - 1
                break
        
        print(f"[INFO] 数据行范围: {data_start_row} 到 {data_end_row}")
        
        # 重新排序序列号
        serial_number = 1
        for row in range(data_start_row, data_end_row + 1):
            # 检查该行是否有设备名称（第2列）
            device_name = ws.cell(row=row, column=2).value
            if device_name and str(device_name).strip():
                # 设置序列号
                ws.cell(row=row, column=1, value=serial_number)
                print(f"  第{row}行: 序号 {serial_number} - {str(device_name)[:30]}...")
                serial_number += 1
            else:
                print(f"  第{row}行: 跳过（无设备名称）")
        
        # 更新合计行的公式（如果存在）
        total_row = data_end_row + 1
        if total_row <= ws.max_row:
            total_cell = ws.cell(row=total_row, column=8)  # H列的合计
            if total_cell.value and '=SUM(' in str(total_cell.value):
                # 更新SUM公式的范围
                new_formula = f"=SUM(H{data_start_row}:H{data_end_row})"
                total_cell.value = new_formula
                print(f"[INFO] 更新合计公式: {new_formula}")
        
        # 保存文件
        output_path = os.path.join(os.path.dirname(__file__), 'templates', '合同清单-视频监控完整内容-最终版.xlsx')
        wb.save(output_path)
        
        print(f"[OK] 序列号修复完成!")
        print(f"[INFO] 输出文件: {output_path}")
        print(f"[INFO] 共处理 {serial_number - 1} 个设备")
        
        return output_path
        
    except Exception as e:
        print(f"[ERROR] 修复失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def verify_file_structure():
    """验证文件结构和内容"""
    file_path = os.path.join(os.path.dirname(__file__), 'templates', '合同清单-视频监控完整内容-最终版.xlsx')
    
    if not os.path.exists(file_path):
        print("[ERROR] 最终版文件不存在")
        return
    
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        print(f"\n[INFO] 验证最终版文件结构:")
        print(f"  - 工作表: {ws.title}")
        print(f"  - 行数: {ws.max_row}")
        print(f"  - 列数: {ws.max_column}")
        
        # 显示表头
        print(f"\n[INFO] 表头信息 (第3行):")
        headers = []
        for col in range(1, 10):  # 假设9列
            header_value = ws.cell(row=3, column=col).value
            if header_value:
                headers.append(str(header_value))
                print(f"  {chr(64+col)}列: {header_value}")
        
        # 显示前5个设备
        print(f"\n[INFO] 前5个设备:")
        for row in range(4, min(9, ws.max_row + 1)):
            serial_num = ws.cell(row=row, column=1).value
            device_name = ws.cell(row=row, column=2).value
            quantity = ws.cell(row=row, column=6).value
            unit_price = ws.cell(row=row, column=7).value
            
            if device_name:
                print(f"  {serial_num}. {str(device_name)[:25]:<25} 数量:{quantity} 单价:{unit_price}")
        
        # 检查合计行
        for row in range(ws.max_row - 3, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and '合计' in str(cell_value):
                total_formula = ws.cell(row=row, column=8).value
                print(f"\n[INFO] 合计行 (第{row}行): {total_formula}")
                break
        
    except Exception as e:
        print(f"[ERROR] 验证失败: {str(e)}")

if __name__ == "__main__":
    print("="*60)
    print("修复Excel文件序列号工具")
    print("="*60)
    
    # 修复序列号
    result = fix_serial_numbers()
    
    if result:
        # 验证文件结构
        verify_file_structure()
        
        print(f"\n[SUCCESS] 修复完成!")
        print(f"最终文件: {result}")
        print("\n修复内容:")
        print("- 序列号从1开始重新排序")
        print("- 跳过无效的空行")
        print("- 更新了合计公式范围")
        print("- 保持了所有格式和样式")
    else:
        print("[FAILED] 修复失败，请检查文件路径和内容")